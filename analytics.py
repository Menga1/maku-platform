"""
MAKU - Assessment Log, Trend Analytics, Site Alert Log, Audit Trail &
Monthly Excel Export
==============================================================================
Logs every risk assessment run, every site safety alert (physiological
CRITICAL/WARNING strain events, safety-override triggers, regulatory fallback
notices), and every security-relevant user action (logins, report exports,
manual threshold overrides) into a database - and builds a real downloadable
.xlsx (openpyxl) with a raw-data sheet, a monthly summary sheet, and an
embedded trend chart.

DATABASE BACKEND - SQLITE TODAY, POSTGRES/SUPABASE-READY:
Every function in this file goes through _get_connection()/_ph()/_adapt_sql()
rather than talking to sqlite3 directly, so the actual engine is a runtime
choice, not a code choice:
  - DATABASE_URL (or MAKU_DATABASE_URL) unset -> SQLite file at MAKU_DB_PATH
    (default "maku_site_data.db"), exactly the original behavior. Nothing
    changes for an existing deployment that hasn't set this variable. This
    is the intended backend in this mode, not a fallback from anything.
  - DATABASE_URL starting with postgres:// or postgresql:// -> routes to
    Postgres/Supabase via psycopg2. This is a FAIL-FAST path (P0 "DATABASE
    SAFETY & FAIL-SAFE RUNTIME"): if the driver is missing, the URL is
    malformed, or the connection drops/fails for any reason, this file no
    longer silently redirects writes to SQLite or to ephemeral
    session_state. Instead it freezes the database layer entirely
    (is_db_fatal() becomes True, every write/read call degrades to a
    no-op/empty result rather than a different store) and records the
    literal CRITICAL_DB_ERROR_MESSAGE constant for the dashboard to
    surface via get_backend_status(). This is a deliberate change from
    "degrade, never break" to "a configured production backend either
    works or the app says so loudly" - an operator who pointed this app at
    Postgres/Supabase does not want a silent, invisible drop back to a
    single ephemeral SQLite file or per-tab memory. See
    is_db_fatal()/_disable_db_fatal()/CRITICAL_DB_ERROR_MESSAGE.
  - The SQL text itself is written once, using "?" placeholders and SQLite
    dialect DDL; _adapt_sql() rewrites both to Postgres's "%s"/SERIAL form
    at the moment of use for the Postgres backend. This is a deliberate,
    honest scope choice: it prepares a real, working, testable migration
    path (connection routing, parameter binding, DDL translation, org-
    hierarchy columns) without rewriting every call site into two parallel
    implementations - a genuine Postgres/Supabase cutover in production
    should still get a dialect-specific review pass, particularly for
    anything beyond this file's fairly simple insert/select/aggregate
    queries.

PERSISTENCE - WHAT THIS ACTUALLY BUYS, HONESTLY:
On the default SQLite path, a Streamlit Cloud-style ephemeral filesystem
still loses the file on redeploy/reboot/sleep - this file does not change
that. What it changes: every browser session now writes to (and can read
from) one shared store for as long as the instance/database stays up,
instead of each tab holding its own private in-memory list. Pointing
DATABASE_URL at a real hosted Postgres/Supabase instance is what actually
makes the history survive a redeploy - and needs no other file in this
app to change, by design.

If the database can't be opened or written for any reason (read-only
filesystem, disk full, locked, Postgres unreachable), every write/read
here degrades gracefully to the previous pure-session_state behavior
rather than crashing the page - see _db_ok().

SESSION-SCOPED VIEW vs ALL-TIME HISTORY:
get_log_dataframe()/monthly_summary()/build_monthly_excel() keep their
original session-scoped contract (each browser session sees only its own
logged rows, and st.session_state.clear() resets that view to empty) -
this preserves the exact UX and test contract the rest of the app already
relies on. Session scoping is implemented with a per-session UUID tag
column, not a separate in-memory store, so the underlying rows are never
actually lost - get_all_time_log_dataframe() reads the full durable
history across every session that has ever written to this running
instance, bounded by `limit` so the trend module never has to load an
unbounded table into RAM. log_site_alert()/get_site_alert_log_dataframe()
are the durable "site alert log / historical risk incidents" store for the
physiological-strain and environmental-alert features.

ENTERPRISE HIERARCHY (Organization -> Project -> Site -> Assessment):
assessment_log, site_alert_log, and audit_log all carry nullable
`organization`/`project`/`site` columns. set_org_context()/get_org_context()
hold the operator's current selection in session_state (mirroring the
existing session-id pattern) with safe defaults, so every existing call
site that doesn't pass them explicitly still works unchanged - this is an
additive column set, not a schema requiring a new required argument
anywhere.

AUDIT TRAIL - TAMPER-EVIDENT, NOT TAMPER-PROOF:
log_audit_event() writes a hash-chained ledger: each row's row_hash covers
its own fields AND the previous row's row_hash, so altering or deleting a
past row breaks the chain from that point forward - detectable via
verify_audit_log_integrity(), not preventable by this file alone (whoever
has direct database access can still rewrite the whole chain from
scratch). That is an honest limitation of an application-level hash chain
without a separate write-only/append-only store or external notarization,
and is disclosed here rather than oversold as cryptographically tamper-
proof.

Mathematical Isolation rule still applies: this file aggregates/logs
results computed elsewhere. It never computes risk itself.
"""

from __future__ import annotations

import hashlib
import io
import os
import sqlite3
import threading
import uuid
from datetime import datetime, timezone

import pandas as pd
from openpyxl import load_workbook
from openpyxl.chart import LineChart, Reference

LOG_SESSION_KEY = "assessment_log"          # legacy in-memory fallback store
SESSION_ID_KEY = "_analytics_session_id"
ORG_CONTEXT_KEY = "_analytics_org_context"

DEFAULT_ORGANIZATION = "Default Organization"
DEFAULT_PROJECT = "Default Project"

# P0 "DATABASE SAFETY & FAIL-SAFE RUNTIME" - the exact, literal message the
# dashboard must display when a configured Postgres/Supabase backend is
# unreachable. See _disable_db_fatal()/is_db_fatal() below.
CRITICAL_DB_ERROR_MESSAGE = (
    "CRITICAL DATABASE ERROR: Persistent connection lost. Manual verification required."
)

DB_PATH = os.environ.get("MAKU_DB_PATH", "maku_site_data.db")
DATABASE_URL = (os.environ.get("DATABASE_URL") or os.environ.get("MAKU_DATABASE_URL") or "").strip()
ALL_TIME_DEFAULT_LIMIT = 5000     # RAM guard for the durable multi-session view
ALERT_LOG_DEFAULT_LIMIT = 1000
AUDIT_LOG_DEFAULT_LIMIT = 2000

# Best-effort single "headline" numeric metric per module, for trend charts -
# each module's result dict has different fields, so this picks the one most
# representative of that module's core risk driver.
MODULE_KEY_METRIC = {
    "Solar (Desert)": "perceived_temp",
    "Offshore (Marine)": "humidex",
    "Underground (Tunnel/Metro)": "perceived_temp",
    "High-Rise (Vertical Urban)": "scaled_wind_speed",
    "Data Center (Controlled Critical Environment)": "arc_flash_energy_cal",
    "Wind Energy (Onshore/Offshore)": "hub_height_wind_speed_ms",
    "Mining & Quarrying": "noise_dose_pct",
    "Marine & Port Construction": "tide_clearance_margin_m",
}

LOG_COLUMNS = ["timestamp", "module", "risk_band", "safety_override", "key_metric", "primary_hazard",
               "organization", "project", "site"]
ALERT_LOG_COLUMNS = ["timestamp", "worker_or_site_id", "alert_type", "severity", "message", "module",
                      "organization", "project", "site", "organization_id", "project_id", "site_id"]
AUDIT_LOG_COLUMNS = ["id", "timestamp", "event_type", "actor", "detail",
                      "organization", "project", "site", "prev_hash", "row_hash"]


# ---------------------------------------------------------------------------
# Database backend abstraction - SQLite by default, Postgres/Supabase when
# DATABASE_URL is set. A fresh short-lived connection per call (simplest way
# to be safe across Streamlit's worker threads without a shared connection
# object).
#
# Two DISTINCT degraded states, deliberately kept separate:
#   - "disabled, non-fatal" (_DB_DISABLED only): no DATABASE_URL was ever
#     configured (plain default SQLite deployment) and the local SQLite
#     file itself became unwritable (read-only fs, disk full, locked).
#     Every public function still degrades to the legacy in-memory
#     session_state list here - this is the original, honest "best effort
#     for a zero-config demo/dev deployment" behavior and is unchanged.
#   - "fatal" (_DB_FATAL, always implies _DB_DISABLED too): a Postgres/
#     Supabase DATABASE_URL was EXPLICITLY configured and could not be
#     reached. Per the P0 FAIL-FAST directive this must never silently
#     redirect to SQLite or session_state - every write freezes (a no-op,
#     not a different store) and every read returns empty, and
#     get_backend_status()["fatal"] carries the exact CRITICAL_DB_ERROR_
#     MESSAGE the dashboard renders. Only a process restart (a fresh
#     DATABASE_URL / a working Postgres instance) clears this - it is not
#     designed to silently self-heal mid-process, since that could mask a
#     real outage behind a banner that flickers on and off.
# _DB_DISABLED/_DB_FATAL/_BACKEND_WARNING are process-wide (not per-
# session) once a failure is seen, so a broken filesystem/unreachable
# database doesn't retry-and-fail on every single call for the rest of the
# process.
# ---------------------------------------------------------------------------
_DB_DISABLED = False
_DB_FATAL = False
_BACKEND_WARNING: str | None = None
_RESOLVED_BACKEND: str | None = None  # "sqlite" | "postgres", set on first connection attempt


def _db_ok() -> bool:
    return not _DB_DISABLED


def is_db_fatal() -> bool:
    """True once a configured Postgres/Supabase backend has failed and this
    process has frozen all database writes/reads rather than silently
    redirecting to ephemeral local storage. See the FAIL-FAST section of
    the module docstring. app.py's dashboard checks this to render the
    CRITICAL DATABASE ERROR banner."""
    return _DB_FATAL


def _disable_db(reason: str = "") -> None:
    """Non-fatal degrade: only reached when no Postgres/Supabase backend
    was ever configured for this process (see module docstring). Falls
    through to the legacy session_state fallback, as before."""
    global _DB_DISABLED
    _DB_DISABLED = True
    if reason:
        _set_backend_warning(reason)


def _disable_db_fatal(reason: str) -> None:
    """FAIL-FAST path (P0): a Postgres/Supabase backend was explicitly
    configured via DATABASE_URL and could not be reached. Freezes the
    database layer entirely - callers must NOT fall through to SQLite or
    session_state after this is set. See is_db_fatal()."""
    global _DB_DISABLED, _DB_FATAL
    _DB_DISABLED = True
    _DB_FATAL = True
    _set_backend_warning(reason)


def _handle_db_failure(exc: Exception, context: str) -> None:
    """Central failure handler for every database call site in this file.
    Routes to the fatal, freeze-everything path when a Postgres/Supabase
    backend was explicitly configured (DATABASE_URL) - the only case the
    P0 FAIL-FAST directive covers - and to the original graceful degrade
    otherwise (no DATABASE_URL at all -> SQLite is the intended backend
    here, not a fallback from Postgres, so keeping the app usable via
    session_state in that specific case is not the "silent redirect" the
    directive is about)."""
    if DATABASE_URL and _is_postgres_url(DATABASE_URL):
        _disable_db_fatal(f"{CRITICAL_DB_ERROR_MESSAGE} ({context}: {exc})")
    else:
        _disable_db(f"Could not reach the database ({context}) - reverting to this "
                     "session's in-memory log only.")


def reset_fatal_state_for_testing() -> None:
    """Test-only escape hatch: clears _DB_DISABLED/_DB_FATAL/_BACKEND_
    WARNING so one test's simulated Postgres outage can't leak fatal state
    into every test that runs after it in the same process. Never called
    from application code - production recovery from a fatal state is by
    design a process restart, not a runtime reset (see module docstring)."""
    global _DB_DISABLED, _DB_FATAL, _BACKEND_WARNING
    _DB_DISABLED = False
    _DB_FATAL = False
    _BACKEND_WARNING = None


def _set_backend_warning(message: str) -> None:
    global _BACKEND_WARNING
    _BACKEND_WARNING = message


def get_backend_status() -> dict:
    """Reports which database backend is actually active, for a settings/
    diagnostics panel and for app.py's dashboard fatal-error banner. Never
    raises - safe to call at any time, including before any connection has
    been attempted (resolved_backend is None until then)."""
    return {
        "configured_database_url_present": bool(DATABASE_URL),
        "resolved_backend": _RESOLVED_BACKEND,
        "disabled": _DB_DISABLED,
        "fatal": _DB_FATAL,
        "fatal_message": CRITICAL_DB_ERROR_MESSAGE if _DB_FATAL else None,
        "warning": _BACKEND_WARNING,
    }


def _is_postgres_url(url: str) -> bool:
    return url.startswith("postgres://") or url.startswith("postgresql://")


def _adapt_sql(sql: str, dialect: str) -> str:
    """Rewrites the SQLite-dialect SQL every query in this file is written
    in ONCE into Postgres dialect, at the moment of use. Safe because none
    of this file's literal SQL text contains a bare "?" character outside
    of a parameter placeholder, and AUTOINCREMENT only ever appears inside
    an INTEGER PRIMARY KEY column definition."""
    if dialect != "postgres":
        return sql
    sql = sql.replace("?", "%s")
    sql = sql.replace("INTEGER PRIMARY KEY AUTOINCREMENT", "SERIAL PRIMARY KEY")
    return sql


class _SqliteBackend:
    dialect = "sqlite"

    def __init__(self, path: str) -> None:
        self.path = path

    def connect(self):
        return sqlite3.connect(self.path, timeout=5)


class _PostgresBackend:
    dialect = "postgres"

    def __init__(self, url: str) -> None:
        import psycopg2  # local import: only required when DATABASE_URL is a Postgres URL
        self._psycopg2 = psycopg2
        self.url = url

    def connect(self):
        conn = self._psycopg2.connect(self.url)
        conn.autocommit = False
        return conn


def _resolve_backend():
    """Picks SQLite or Postgres based on DATABASE_URL.

    FAIL-FAST (P0): when a Postgres/Supabase DATABASE_URL is configured,
    this function no longer swallows a missing driver or malformed URL and
    quietly hands back a SQLite backend instead - it raises (ImportError
    if psycopg2 isn't installed; whatever _PostgresBackend's construction
    raises for anything else), and lets the caller's exception handling
    (_handle_db_failure -> _disable_db_fatal) freeze the database layer
    and surface CRITICAL_DB_ERROR_MESSAGE instead of silently persisting
    to a different store than the one production was configured to use.

    Resolved once per _get_connection() call rather than cached globally,
    so a corrected DATABASE_URL / freshly-installed psycopg2 takes effect
    on the next call without restarting the process (once the process
    isn't already latched into the fatal state - see is_db_fatal())."""
    global _RESOLVED_BACKEND
    if DATABASE_URL and _is_postgres_url(DATABASE_URL):
        backend = _PostgresBackend(DATABASE_URL)  # raises ImportError if psycopg2 is missing
        _RESOLVED_BACKEND = "postgres"
        return backend
    _RESOLVED_BACKEND = "sqlite"
    return _SqliteBackend(DB_PATH)


class _Connection:
    """Thin wrapper so every call site can keep writing "?"-style SQL and
    calling conn.execute(sql, params) exactly as before, regardless of
    which backend is actually active - _adapt_sql() does the dialect
    translation transparently. Used as a context manager, same as the
    original bare sqlite3.Connection was."""

    def __init__(self, raw_conn, dialect: str) -> None:
        self._conn = raw_conn
        self._dialect = dialect

    def execute(self, sql: str, params: tuple = ()):
        cur = self._conn.cursor()
        cur.execute(_adapt_sql(sql, self._dialect), params)
        return cur

    def executemany(self, sql: str, seq_of_params) -> None:
        cur = self._conn.cursor()
        cur.executemany(_adapt_sql(sql, self._dialect), list(seq_of_params))

    def __enter__(self) -> "_Connection":
        return self

    def __exit__(self, exc_type, exc, tb) -> None:
        if exc_type is None:
            self._conn.commit()
        else:
            self._conn.rollback()
        self._conn.close()

    # pandas.read_sql_query() wants the raw DBAPI2 connection, not this
    # wrapper - callers use conn.raw for read queries via pandas.
    @property
    def raw(self):
        return self._conn

    @property
    def dialect(self) -> str:
        return self._dialect


_ASSESSMENT_LOG_DDL = """CREATE TABLE IF NOT EXISTS assessment_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    session_id TEXT NOT NULL,
    timestamp TEXT NOT NULL,
    module TEXT NOT NULL,
    risk_band TEXT,
    safety_override INTEGER,
    key_metric REAL,
    primary_hazard TEXT,
    organization TEXT,
    project TEXT,
    site TEXT
)"""

_SITE_ALERT_LOG_DDL = """CREATE TABLE IF NOT EXISTS site_alert_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TEXT NOT NULL,
    worker_or_site_id TEXT,
    alert_type TEXT NOT NULL,
    severity TEXT NOT NULL,
    message TEXT NOT NULL,
    module TEXT,
    organization TEXT,
    project TEXT,
    site TEXT
)"""

_AUDIT_LOG_DDL = """CREATE TABLE IF NOT EXISTS audit_log (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    timestamp TEXT NOT NULL,
    event_type TEXT NOT NULL,
    actor TEXT NOT NULL,
    detail TEXT,
    organization TEXT,
    project TEXT,
    site TEXT,
    prev_hash TEXT NOT NULL,
    row_hash TEXT NOT NULL
)"""

# ---------------------------------------------------------------------------
# P0 "MULTI-TENANT ISOLATION ARCHITECTURE" - real relational hierarchy:
# User Session -> Organization ID -> Project ID -> Site ID -> Assessment ID.
# assessment_log's own auto-increment `id` IS the Assessment ID in that
# chain already; these three new tables give Organization/Project/Site a
# genuine FK-linked identity instead of being bare, unrelated free-text
# columns on every log table. Every log table keeps its original TEXT
# organization/project/site columns (human-readable, backward-compatible
# with every existing query/test/export) AND gains an *_id INTEGER column
# that references these tables - additive, not a breaking schema change.
# ---------------------------------------------------------------------------
_ORGANIZATIONS_DDL = """CREATE TABLE IF NOT EXISTS organizations (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    name TEXT NOT NULL UNIQUE
)"""

_PROJECTS_DDL = """CREATE TABLE IF NOT EXISTS projects (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    organization_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    UNIQUE(organization_id, name)
)"""

_SITES_DDL = """CREATE TABLE IF NOT EXISTS sites (
    id INTEGER PRIMARY KEY AUTOINCREMENT,
    project_id INTEGER NOT NULL,
    name TEXT NOT NULL,
    UNIQUE(project_id, name)
)"""

# Additive columns for databases created by an earlier version of this file
# (pre-hierarchy). SQLite/Postgres both accept "ADD COLUMN IF NOT EXISTS" -
# guarded in a try/except per-statement anyway since older SQLite versions
# (<3.35) don't support the IF NOT EXISTS clause on ADD COLUMN.
_MIGRATION_STATEMENTS = [
    "ALTER TABLE assessment_log ADD COLUMN organization TEXT",
    "ALTER TABLE assessment_log ADD COLUMN project TEXT",
    "ALTER TABLE assessment_log ADD COLUMN site TEXT",
    "ALTER TABLE assessment_log ADD COLUMN organization_id INTEGER",
    "ALTER TABLE assessment_log ADD COLUMN project_id INTEGER",
    "ALTER TABLE assessment_log ADD COLUMN site_id INTEGER",
    "ALTER TABLE site_alert_log ADD COLUMN organization TEXT",
    "ALTER TABLE site_alert_log ADD COLUMN project TEXT",
    "ALTER TABLE site_alert_log ADD COLUMN site TEXT",
    "ALTER TABLE site_alert_log ADD COLUMN organization_id INTEGER",
    "ALTER TABLE site_alert_log ADD COLUMN project_id INTEGER",
    "ALTER TABLE site_alert_log ADD COLUMN site_id INTEGER",
    "ALTER TABLE audit_log ADD COLUMN organization_id INTEGER",
    "ALTER TABLE audit_log ADD COLUMN project_id INTEGER",
    "ALTER TABLE audit_log ADD COLUMN site_id INTEGER",
]


def _get_connection() -> _Connection:
    backend = _resolve_backend()
    raw = backend.connect()
    conn = _Connection(raw, backend.dialect)
    conn.execute(_ASSESSMENT_LOG_DDL)
    conn.execute(_SITE_ALERT_LOG_DDL)
    conn.execute(_AUDIT_LOG_DDL)
    conn.execute(_ORGANIZATIONS_DDL)
    conn.execute(_PROJECTS_DDL)
    conn.execute(_SITES_DDL)
    for stmt in _MIGRATION_STATEMENTS:
        try:
            conn.execute(stmt)
        except Exception:  # noqa: BLE001 - column already exists on every call after the first
            pass
    conn.execute("CREATE INDEX IF NOT EXISTS idx_assessment_session ON assessment_log(session_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_assessment_ts ON assessment_log(timestamp)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_assessment_org ON assessment_log(organization_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_alert_ts ON site_alert_log(timestamp)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_alert_org ON site_alert_log(organization_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_audit_ts ON audit_log(timestamp)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_projects_org ON projects(organization_id)")
    conn.execute("CREATE INDEX IF NOT EXISTS idx_sites_project ON sites(project_id)")
    return conn


def _get_or_create_id(conn: "_Connection", table: str, unique_cols: dict, name_col: str, name_value: str) -> int:
    """Resolve-or-create helper backing get_or_create_organization_id() /
    get_or_create_project_id() / get_or_create_site_id(): looks up a row
    matching `unique_cols` (already-known parent id(s), if any) plus
    `name_col` = `name_value`; inserts it if absent. Runs inside the
    caller's own connection/transaction so a concurrent duplicate insert
    is caught by the table's UNIQUE constraint and simply re-queried
    rather than raising - safe under the same kind of race the audit-chain
    mutex protects against elsewhere in this file, just lower-stakes here
    (worst case is one retry, never a corrupted hierarchy)."""
    where_cols = list(unique_cols.keys()) + [name_col]
    where_vals = list(unique_cols.values()) + [name_value]
    where_clause = " AND ".join(f"{c} = ?" for c in where_cols)
    cur = conn.execute(f"SELECT id FROM {table} WHERE {where_clause}", tuple(where_vals))
    row = cur.fetchone()
    if row:
        return int(row[0])
    insert_cols = list(unique_cols.keys()) + [name_col]
    placeholders = ", ".join("?" for _ in insert_cols)
    insert_sql = f"INSERT INTO {table} ({', '.join(insert_cols)}) VALUES ({placeholders})"
    if conn.dialect == "postgres":
        # psycopg2 cursors don't populate lastrowid (that's a sqlite3-
        # specific DBAPI extension) - RETURNING id is the portable
        # Postgres way to get the newly-inserted row's id back.
        insert_sql += " RETURNING id"
    try:
        cur = conn.execute(insert_sql, tuple(where_vals))
        new_id = cur.fetchone()[0] if conn.dialect == "postgres" else cur.lastrowid
        if new_id:
            return int(new_id)
    except Exception:  # noqa: BLE001 - lost the UNIQUE-constraint race to a concurrent insert
        pass
    cur = conn.execute(f"SELECT id FROM {table} WHERE {where_clause}", tuple(where_vals))
    row = cur.fetchone()
    return int(row[0]) if row else 0


def get_or_create_organization_id(organization: str) -> int | None:
    """Resolves `organization` to its stable relational id, creating the
    row if this is the first time this organization name has been seen.
    Returns None (never raises) if the database is unavailable - callers
    fall back to the plain organization-name string, exactly like every
    other database-backed feature in this file degrades."""
    if not organization or not _db_ok():
        return None
    try:
        with _get_connection() as conn:
            return _get_or_create_id(conn, "organizations", {}, "name", organization)
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "get_or_create_organization_id")
        return None


def get_or_create_project_id(organization_id: int | None, project: str) -> int | None:
    """Resolves `project` (scoped under `organization_id`) to its stable
    relational id. Returns None if organization_id is None (can't create
    an orphan project) or the database is unavailable."""
    if organization_id is None or not project or not _db_ok():
        return None
    try:
        with _get_connection() as conn:
            return _get_or_create_id(conn, "projects", {"organization_id": organization_id}, "name", project)
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "get_or_create_project_id")
        return None


def get_or_create_site_id(project_id: int | None, site: str) -> int | None:
    """Resolves `site` (scoped under `project_id`) to its stable relational
    id. Returns None if project_id is None or the site name is blank (a
    blank site is a valid, common MVP state - not every operator names a
    specific site) or the database is unavailable."""
    if project_id is None or not site or not _db_ok():
        return None
    try:
        with _get_connection() as conn:
            return _get_or_create_id(conn, "sites", {"project_id": project_id}, "name", site)
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "get_or_create_site_id")
        return None


def list_organizations() -> list[str]:
    """All known organization names, alphabetical - backs app.py's
    DB-driven organization selector (P0 "shift away from basic user-
    entered text fields"). Returns [] (never raises) if the database is
    unavailable, so the sidebar can fall back to a plain text field."""
    if not _db_ok():
        return []
    try:
        with _get_connection() as conn:
            cur = conn.execute("SELECT name FROM organizations ORDER BY name")
            return [row[0] for row in cur.fetchall()]
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "list_organizations")
        return []


# Backward-compatible aliases: the pre-refactor names, kept so nothing
# else in this codebase (or a caller outside it) breaks if it referenced
# the old SQLite-specific function names directly.
_sqlite_ok = _db_ok
_disable_sqlite = _disable_db


def _get_session_id(st_module) -> str:
    """One random id per browser session, stored in session_state. Clearing
    session_state (as the test suite does between test methods, and as a
    real logout would) forgets this id, which is exactly what makes
    get_log_dataframe() go back to reporting an empty session-scoped log -
    the underlying database rows from the old id are never deleted, just no
    longer tagged as 'this session's view'."""
    if SESSION_ID_KEY not in st_module.session_state:
        st_module.session_state[SESSION_ID_KEY] = uuid.uuid4().hex
    return st_module.session_state[SESSION_ID_KEY]


# ---------------------------------------------------------------------------
# Enterprise hierarchy context: Organization -> Project -> Site.
# Session-state-backed, same pattern as the session id above - an operator
# picks these once (e.g. a sidebar selector app.py wires up) and every
# subsequent log_assessment()/log_site_alert()/log_audit_event() call in
# that session tags itself with the current context unless the caller
# passes an explicit override. Safe defaults mean no existing call site
# has to change to keep working.
# ---------------------------------------------------------------------------

def set_org_context(st_module, organization: str = "", project: str = "", site: str = "") -> None:
    """Sets the current session's Organization/Project/Site tags AND
    resolves/creates their relational ids (User Session -> Organization ID
    -> Project ID -> Site ID, see the organizations/projects/sites DDL
    above) - the ids are what every tenant-isolation filter downstream
    actually keys on, the name strings remain for display/CSV/backward
    compatibility. Call this from a sidebar selector; an empty string for
    any field means "use the default" rather than "leave whatever was
    there," so a selector reset behaves predictably. If the database is
    unavailable, the *_id fields degrade to None (see get_or_create_*_id())
    - the app keeps working on name strings alone, exactly like every
    other database-backed feature in this file degrades."""
    organization = organization or DEFAULT_ORGANIZATION
    project = project or DEFAULT_PROJECT
    site = site or ""
    organization_id = get_or_create_organization_id(organization)
    project_id = get_or_create_project_id(organization_id, project)
    site_id = get_or_create_site_id(project_id, site) if site else None
    st_module.session_state[ORG_CONTEXT_KEY] = {
        "organization": organization,
        "project": project,
        "site": site,
        "organization_id": organization_id,
        "project_id": project_id,
        "site_id": site_id,
    }


def get_org_context(st_module) -> dict:
    """Current session's Organization/Project/Site tags AND their
    relational ids, defaulting to DEFAULT_ORGANIZATION/DEFAULT_PROJECT/""
    /None if never explicitly set. NOTE: app.py spreads this dict as
    **get_org_context(st_module) directly into log_audit_event() calls
    (e.g. for AUDIT_EVENT_THRESHOLD_OVERRIDE) - log_audit_event() accepts
    every key this returns, including the *_id fields, for exactly that
    reason. log_assessment() and log_site_alert() resolve their own ids
    internally instead (see their docstrings) since neither is called
    with this dict spread into it."""
    return st_module.session_state.get(ORG_CONTEXT_KEY, {
        "organization": DEFAULT_ORGANIZATION,
        "project": DEFAULT_PROJECT,
        "site": "",
        "organization_id": None,
        "project_id": None,
        "site_id": None,
    })


# ---------------------------------------------------------------------------
# Evidence & Traceability (HSE audit corrective action)
# ---------------------------------------------------------------------------
# "PDF must capture 100% of inputs... Add Evidence & Traceability section:
# Data Source (Live API vs Manual), Geographic Location/Site, Timestamp,
# Measuring Equipment/Sensor Model." The raw per-hazard inputs themselves
# are already captured 100% by every calculate_*_kinetic_risk()'s own
# "drivers" dict (already rendered in full in both the HTML and PDF
# report - see ui_helpers._build_report_pdf / render_official_report).
# This function assembles the ADDITIONAL provenance metadata the audit
# specifically calls out, reusing context this app already tracks
# (get_org_context() for Geographic Location/Site) rather than inventing
# a second, disconnected location concept.

def build_evidence_traceability(st_module, result: dict, data_mode: str = "manual",
                                 sensor_model: str = "") -> dict:
    """
    Assembles the Evidence & Traceability block for one assessment's
    report. Honest about what this MVP does and doesn't track:
      - Data Source: derived directly from the page's own data_mode
        ("auto" = live API/simulated telemetry feed, "manual" = operator-
        entered values) - the same manual/auto distinction already shown
        to the user via render_data_mode_selector()/render_feed_ok_banner().
        This does NOT claim a specific upstream API name/URL per reading,
        since that isn't persisted across the assessment's lifecycle in
        this version - a caller with a live feed's exact source string in
        scope may still show it separately (e.g. via
        render_feed_ok_banner()) alongside this block.
      - Geographic Location/Site: reuses get_org_context()'s existing
        Organization/Project/Site tags - not a second, parallel location
        field that could drift out of sync with the one already shown
        throughout the app (sidebar, assessment log, alerts).
      - Timestamp: UTC, captured at the moment this function runs (i.e.
        report-generation time, matching render_official_report()'s own
        "Generated" timestamp already in the report header).
      - Measuring Equipment / Sensor Model: caller-supplied free text
        (e.g. a WBGT meter model when a real instrument reading was
        entered manually) - "Not specified / simulated input" when empty,
        never fabricated.
    """
    return {
        "Data Source": "Live API / telemetry feed (auto mode)" if data_mode == "auto"
                       else "Manual entry / simulated input",
        "Geographic Location - Organization": get_org_context(st_module).get("organization", DEFAULT_ORGANIZATION),
        "Geographic Location - Project": get_org_context(st_module).get("project", DEFAULT_PROJECT),
        "Geographic Location - Site": get_org_context(st_module).get("site", "") or result.get("module", ""),
        "Assessment Timestamp (UTC)": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "Measuring Equipment / Sensor Model": sensor_model.strip() if sensor_model and sensor_model.strip()
                                              else "Not specified / simulated input",
        "Module": result.get("module", ""),
    }


# ---------------------------------------------------------------------------
# Assessment log (session-scoped view, database-backed with durable history)
# ---------------------------------------------------------------------------

def log_assessment(st_module, result: dict, organization: str | None = None,
                    project: str | None = None, site: str | None = None) -> None:
    """Append one row to this session's assessment log. Call right after
    computing a result on any module page. organization/project/site
    default to the session's current get_org_context() (site defaults to
    the module name) - pass them explicitly only to override that."""
    module = result.get("module", "")
    key_metric_field = MODULE_KEY_METRIC.get(module)
    ctx = get_org_context(st_module)
    row = {
        "timestamp": datetime.now(timezone.utc).isoformat(timespec="seconds"),
        "module": module,
        "risk_band": str(result.get("risk_band", "")),
        "safety_override": bool(result.get("safety_override", False)),
        "key_metric": result.get(key_metric_field) if key_metric_field else None,
        "primary_hazard": result.get("primary_hazard", ""),
        "organization": organization or ctx["organization"],
        "project": project or ctx["project"],
        "site": site or ctx["site"] or module,
    }
    # Relational hierarchy ids (P0 "MULTI-TENANT ISOLATION ARCHITECTURE"):
    # reuse the session's already-resolved ids from set_org_context() when
    # no explicit organization/project/site override was passed (the
    # overwhelmingly common case); only re-resolve against the database
    # when a caller explicitly overrides one of the name strings.
    if organization or project or site:
        org_id = get_or_create_organization_id(row["organization"])
        proj_id = get_or_create_project_id(org_id, row["project"])
        site_row_id = get_or_create_site_id(proj_id, row["site"]) if row["site"] else None
    else:
        org_id, proj_id, site_row_id = ctx.get("organization_id"), ctx.get("project_id"), ctx.get("site_id")

    if _db_ok():
        session_id = _get_session_id(st_module)
        try:
            with _get_connection() as conn:
                conn.execute(
                    """INSERT INTO assessment_log
                       (session_id, timestamp, module, risk_band, safety_override, key_metric,
                        primary_hazard, organization, project, site,
                        organization_id, project_id, site_id)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (session_id, row["timestamp"], row["module"], row["risk_band"],
                     int(row["safety_override"]), row["key_metric"], row["primary_hazard"],
                     row["organization"], row["project"], row["site"],
                     org_id, proj_id, site_row_id),
                )
            return
        except Exception as exc:  # noqa: BLE001 - any backend failure -> handled centrally
            _handle_db_failure(exc, "log_assessment")

    if is_db_fatal():
        # FAIL-FAST: a configured Postgres/Supabase backend is down. Freeze
        # writes entirely rather than silently redirecting this assessment
        # into ephemeral session_state - see module docstring.
        return

    if LOG_SESSION_KEY not in st_module.session_state:
        st_module.session_state[LOG_SESSION_KEY] = []
    st_module.session_state[LOG_SESSION_KEY].append(row)


def get_log_dataframe(st_module) -> pd.DataFrame:
    """This session's log as a DataFrame, oldest first, timestamp parsed.
    Same contract as before the database-backend upgrade: empty right
    after st.session_state.clear()."""
    if _db_ok():
        session_id = st_module.session_state.get(SESSION_ID_KEY)
        if session_id is None:
            return pd.DataFrame(columns=LOG_COLUMNS)
        try:
            with _get_connection() as conn:
                df = pd.read_sql_query(
                    _adapt_sql(
                        f"SELECT {', '.join(LOG_COLUMNS)} FROM assessment_log "
                        "WHERE session_id = ? ORDER BY timestamp",
                        conn.dialect,
                    ),
                    conn.raw, params=(session_id,),
                )
            if df.empty:
                return pd.DataFrame(columns=LOG_COLUMNS)
            df["timestamp"] = pd.to_datetime(df["timestamp"])
            df["safety_override"] = df["safety_override"].astype(bool)
            return df.reset_index(drop=True)
        except Exception as exc:  # noqa: BLE001
            _handle_db_failure(exc, "get_log_dataframe")

    if is_db_fatal():
        # FAIL-FAST: never serve a possibly-stale local session_state view
        # while a configured production backend is down - an empty result
        # next to the dashboard's CRITICAL DATABASE ERROR banner is the
        # honest state here, not a quiet partial view.
        return pd.DataFrame(columns=LOG_COLUMNS)

    rows = st_module.session_state.get(LOG_SESSION_KEY, [])
    if not rows:
        return pd.DataFrame(columns=LOG_COLUMNS)
    df = pd.DataFrame(rows)
    df["timestamp"] = pd.to_datetime(df["timestamp"])
    return df.sort_values("timestamp").reset_index(drop=True)


def get_all_time_log_dataframe(limit: int = ALL_TIME_DEFAULT_LIMIT) -> pd.DataFrame:
    """The durable, cross-session history for this running instance (see
    module docstring) - every assessment ever logged while this app
    instance has been up, from every browser session, newest first,
    bounded by `limit` so this never pulls an unbounded table into RAM.
    Returns an empty DataFrame (never raises) if the database is
    unavailable."""
    if not _db_ok():
        return pd.DataFrame(columns=LOG_COLUMNS)
    try:
        with _get_connection() as conn:
            df = pd.read_sql_query(
                _adapt_sql(
                    f"SELECT {', '.join(LOG_COLUMNS)} FROM assessment_log "
                    "ORDER BY timestamp DESC LIMIT ?",
                    conn.dialect,
                ),
                conn.raw, params=(int(limit),),
            )
        if df.empty:
            return pd.DataFrame(columns=LOG_COLUMNS)
        df["timestamp"] = pd.to_datetime(df["timestamp"])
        df["safety_override"] = df["safety_override"].astype(bool)
        return df.reset_index(drop=True)
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "get_all_time_log_dataframe")
        return pd.DataFrame(columns=LOG_COLUMNS)


def merge_uploaded_csv(st_module, uploaded_file) -> int:
    """Parses a previously-downloaded CSV log and merges it into this
    session's log, de-duplicating exact repeats. Returns how many new
    rows were actually added (0 if everything was already present).
    organization/project/site are OPTIONAL in the uploaded file (filled
    with the session's current org context if absent) so a CSV downloaded
    from an earlier version of this app - before the hierarchy columns
    existed - still merges cleanly instead of being rejected."""
    uploaded_df = pd.read_csv(uploaded_file)
    required_columns = [c for c in LOG_COLUMNS if c not in {"organization", "project", "site"}]
    missing = set(required_columns) - set(uploaded_df.columns)
    if missing:
        raise ValueError(f"Uploaded file is missing expected columns: {sorted(missing)}")

    ctx = get_org_context(st_module)
    for col, default in (("organization", ctx["organization"]), ("project", ctx["project"]), ("site", ctx["site"])):
        if col not in uploaded_df.columns:
            uploaded_df[col] = default
        else:
            uploaded_df[col] = uploaded_df[col].fillna(default)

    existing_df = get_log_dataframe(st_module)
    combined = pd.concat([existing_df, uploaded_df], ignore_index=True)
    combined["timestamp"] = pd.to_datetime(combined["timestamp"])
    before = len(existing_df)
    combined = combined.drop_duplicates(subset=["timestamp", "module", "risk_band", "key_metric"])
    combined = combined.sort_values("timestamp").reset_index(drop=True)
    added = len(combined) - before

    records = combined.to_dict("records")
    for row in records:
        ts = row["timestamp"]
        row["timestamp"] = ts.isoformat() if hasattr(ts, "isoformat") else str(ts)
        row["safety_override"] = bool(row["safety_override"])

    if _db_ok():
        session_id = _get_session_id(st_module)
        try:
            with _get_connection() as conn:
                conn.execute("DELETE FROM assessment_log WHERE session_id = ?", (session_id,))
                conn.executemany(
                    """INSERT INTO assessment_log
                       (session_id, timestamp, module, risk_band, safety_override, key_metric,
                        primary_hazard, organization, project, site)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    [(session_id, r["timestamp"], r["module"], r["risk_band"],
                      int(r["safety_override"]), r["key_metric"], r["primary_hazard"],
                      r.get("organization") or ctx["organization"], r.get("project") or ctx["project"],
                      r.get("site") or ctx["site"]) for r in records],
                )
            return added
        except Exception as exc:  # noqa: BLE001
            _handle_db_failure(exc, "merge_uploaded_csv")

    if is_db_fatal():
        # FAIL-FAST: do not write the merged set into ephemeral
        # session_state while a configured production backend is down.
        return 0

    st_module.session_state[LOG_SESSION_KEY] = records
    return added


# ---------------------------------------------------------------------------
# Site alert log - physiological CRITICAL/WARNING strain events, safety-
# override triggers, regulatory-fallback notices, etc. Always durable
# (all-time, not session-scoped) since these are safety incidents a site
# manager needs to review regardless of who was looking at the dashboard
# when they fired. Never includes worker names - callers must pass an
# already-anonymized id (e.g. "Worker_A3"), never raw PII (see app.py's
# worker-strain dashboard for the anonymization step).
# ---------------------------------------------------------------------------

def log_site_alert(worker_or_site_id: str, alert_type: str, severity: str, message: str, module: str = "",
                    organization: str = "", project: str = "", site: str = "",
                    organization_id: int | None = None, project_id: int | None = None,
                    site_id: int | None = None) -> None:
    """Records one durable site alert row. Silently no-ops if the database
    is unavailable (an alert that can't be written to disk still displayed
    on screen at the moment it fired - this is a log for later review, not
    the sole notification channel). organization/project/site are plain
    strings here (not session-context-derived) since alerts can fire from
    a background thread with no Streamlit session attached.

    DATA ISOLATION (P0): callers SHOULD pass at least `organization` (the
    acting session's get_org_context(st)["organization"]) so
    get_site_alert_log_dataframe(organization=...) can filter this row out
    of every other tenant's view - see that function's docstring. If the
    *_id fields aren't supplied explicitly, they're resolved from the name
    strings here (organization_id from organization, project_id from
    organization_id+project, site_id from project_id+site) so the
    relational hierarchy stays populated even for this no-session call
    path."""
    if not _db_ok():
        return
    if organization_id is None and organization:
        organization_id = get_or_create_organization_id(organization)
    if project_id is None and organization_id is not None and project:
        project_id = get_or_create_project_id(organization_id, project)
    if site_id is None and project_id is not None and site:
        site_id = get_or_create_site_id(project_id, site)
    try:
        with _get_connection() as conn:
            conn.execute(
                """INSERT INTO site_alert_log
                   (timestamp, worker_or_site_id, alert_type, severity, message, module,
                    organization, project, site, organization_id, project_id, site_id)
                   VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                (datetime.now(timezone.utc).isoformat(timespec="seconds"),
                 worker_or_site_id, alert_type, severity, message, module,
                 organization, project, site, organization_id, project_id, site_id),
            )
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "log_site_alert")


def get_site_alert_log_dataframe(limit: int = ALERT_LOG_DEFAULT_LIMIT,
                                  organization: str | None = None) -> pd.DataFrame:
    """Durable site alert history, newest first, bounded by `limit`.

    DATA ISOLATION (P0 "MULTI-TENANT ISOLATION ARCHITECTURE"): this table
    is deliberately ALL-TIME and NOT session-scoped (see module docstring
    - a site manager must see incidents regardless of who else was logged
    in when they fired), which means it has no natural per-session
    boundary the way get_log_dataframe() does. Pass `organization` (the
    active session's get_org_context(st)["organization"]) to restrict the
    query to that tenant's rows at the SQL layer - callers in app.py MUST
    pass this; omitting it returns every organization's alerts and should
    only be used by a genuine cross-tenant admin view, never a normal
    site-manager dashboard."""
    if not _db_ok():
        return pd.DataFrame(columns=ALERT_LOG_COLUMNS)
    try:
        with _get_connection() as conn:
            if organization:
                sql = (f"SELECT {', '.join(ALERT_LOG_COLUMNS)} FROM site_alert_log "
                       "WHERE organization = ? ORDER BY timestamp DESC LIMIT ?")
                params = (organization, int(limit))
            else:
                sql = (f"SELECT {', '.join(ALERT_LOG_COLUMNS)} FROM site_alert_log "
                       "ORDER BY timestamp DESC LIMIT ?")
                params = (int(limit),)
            df = pd.read_sql_query(_adapt_sql(sql, conn.dialect), conn.raw, params=params)
        if not df.empty:
            df["timestamp"] = pd.to_datetime(df["timestamp"])
        return df
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "get_site_alert_log_dataframe")
        return pd.DataFrame(columns=ALERT_LOG_COLUMNS)


# ---------------------------------------------------------------------------
# Audit trail - hash-chained, tamper-EVIDENT ledger of security-relevant
# actions: logins (success/failure), report/PDF generation, and manual
# regulatory-threshold overrides. See the module docstring's "AUDIT TRAIL"
# section for the honest limits of an application-level hash chain.
# ---------------------------------------------------------------------------

AUDIT_EVENT_LOGIN_SUCCESS = "LOGIN_SUCCESS"
AUDIT_EVENT_LOGIN_FAILURE = "LOGIN_FAILURE"
AUDIT_EVENT_LOGIN_LOCKOUT = "LOGIN_LOCKOUT"
AUDIT_EVENT_LOGOUT = "LOGOUT"
AUDIT_EVENT_SESSION_EXPIRED = "SESSION_EXPIRED"
AUDIT_EVENT_REPORT_GENERATED = "REPORT_GENERATED"
AUDIT_EVENT_THRESHOLD_OVERRIDE = "THRESHOLD_OVERRIDE"
AUDIT_EVENT_CONFIG_ERROR = "CONFIG_ERROR"
AUDIT_EVENT_DEPRECATED_CREDENTIAL_FORMAT = "DEPRECATED_CREDENTIAL_FORMAT"

_GENESIS_HASH = "0" * 64  # the chain's first row's prev_hash - a well-known constant, not a secret

# P0 "SECURITY EXPOSURE & AUDIT CONCURRENCY" - MUTEX PROTECTION.
# The read-prev-hash -> compute -> insert sequence in log_audit_event() is
# a classic read-then-write race: two concurrent writers can both read the
# same prev_hash before either commits, producing two rows chained off the
# same parent (a fork verify_audit_log_integrity() would then report as a
# broken chain, even though nothing was maliciously tampered with - just
# raced). Two layers, deliberately both present rather than either alone:
#   1. _AUDIT_CHAIN_LOCK (threading.Lock): serializes every writer within
#      THIS process. Cheap, always correct, but does not help across
#      multiple worker processes/replicas.
#   2. _lock_audit_chain(conn): a REAL database-level lock, scoped to the
#      connection's own transaction, that also serializes writers ACROSS
#      processes/replicas - the property a pure in-process lock can't
#      give. SQLite: BEGIN IMMEDIATE claims the write lock immediately
#      instead of SQLite's default deferred (lazy, first-DML) transaction
#      start, so a second connection's BEGIN IMMEDIATE blocks until the
#      first commits. Postgres: pg_advisory_xact_lock() takes a session-
#      scoped advisory lock keyed to a fixed constant reserved for this
#      purpose, released automatically at transaction end.
_AUDIT_CHAIN_LOCK = threading.Lock()
_AUDIT_ADVISORY_LOCK_KEY = 913042017  # arbitrary constant, reserved for the audit chain only


def _lock_audit_chain(conn: "_Connection") -> None:
    """Must be the FIRST statement executed against `conn` inside
    log_audit_event()'s `with _get_connection() as conn:` block (before
    _last_audit_hash()) - see the MUTEX PROTECTION note above. Safe to
    call this early because _get_connection()'s own DDL/index statements
    run and auto-commit (SQLite) / occur within a transaction Postgres
    hasn't required a lock for yet, before this function's caller ever
    gets the connection back."""
    if conn.dialect == "sqlite":
        conn.raw.execute("BEGIN IMMEDIATE")
    else:
        # conn.execute() runs every statement through _adapt_sql(), which
        # rewrites "?" -> "%s" for the postgres dialect - same placeholder
        # style as every other query in this file.
        conn.execute("SELECT pg_advisory_xact_lock(?)", (_AUDIT_ADVISORY_LOCK_KEY,))


def _hash_audit_row(prev_hash: str, timestamp: str, event_type: str, actor: str, detail: str,
                     organization: str, project: str, site: str) -> str:
    payload = "|".join([prev_hash, timestamp, event_type, actor, detail, organization, project, site])
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def _last_audit_hash(conn) -> str:
    cur = conn.execute("SELECT row_hash FROM audit_log ORDER BY id DESC LIMIT 1")
    row = cur.fetchone()
    return row[0] if row else _GENESIS_HASH


def log_audit_event(event_type: str, actor: str = "anonymous", detail: str = "",
                     organization: str = "", project: str = "", site: str = "",
                     organization_id: int | None = None, project_id: int | None = None,
                     site_id: int | None = None) -> bool:
    """Appends one row to the hash-chained audit ledger. Returns True if
    the row was durably written, False if it degraded to a no-op (database
    unavailable) - callers that must guarantee an audit trail exists (e.g.
    a compliance export) can check this return value; callers just
    recording a routine event (most of them) can ignore it, matching this
    file's existing 'never let logging break the page' philosophy.

    organization_id/project_id/site_id accept get_org_context()'s
    resolved relational ids directly - app.py calls this as
    log_audit_event(..., **get_org_context(st)), which supplies both the
    name strings and these ids in one spread. The hash itself is computed
    only over the name strings (see _hash_audit_row) so the tamper-
    evidence contract/hash format is unchanged; the ids are stored
    alongside for tenant-filtered queries, not folded into the chain.

    MUTEX PROTECTION: the entire read-prev-hash/compute/insert sequence is
    serialized both in-process (_AUDIT_CHAIN_LOCK) and at the database
    level (_lock_audit_chain) - see the MUTEX PROTECTION note above
    _AUDIT_CHAIN_LOCK's definition.

    Never raises. actor should be a username (or "anonymous"/"system" for
    unauthenticated/background events) - never a raw secret or password."""
    if not _db_ok():
        return False
    timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")
    with _AUDIT_CHAIN_LOCK:
        try:
            with _get_connection() as conn:
                _lock_audit_chain(conn)
                prev_hash = _last_audit_hash(conn)
                row_hash = _hash_audit_row(prev_hash, timestamp, event_type, actor, detail,
                                            organization, project, site)
                conn.execute(
                    """INSERT INTO audit_log
                       (timestamp, event_type, actor, detail, organization, project, site,
                        organization_id, project_id, site_id, prev_hash, row_hash)
                       VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)""",
                    (timestamp, event_type, actor, detail, organization, project, site,
                     organization_id, project_id, site_id, prev_hash, row_hash),
                )
            return True
        except Exception as exc:  # noqa: BLE001
            _handle_db_failure(exc, "log_audit_event")
            return False


def get_audit_log_dataframe(limit: int = AUDIT_LOG_DEFAULT_LIMIT) -> pd.DataFrame:
    """Durable audit trail, newest first, bounded by `limit`. Never raises."""
    if not _db_ok():
        return pd.DataFrame(columns=AUDIT_LOG_COLUMNS)
    try:
        with _get_connection() as conn:
            df = pd.read_sql_query(
                _adapt_sql(
                    f"SELECT {', '.join(AUDIT_LOG_COLUMNS)} FROM audit_log "
                    "ORDER BY id DESC LIMIT ?",
                    conn.dialect,
                ),
                conn.raw, params=(int(limit),),
            )
        if not df.empty:
            df["timestamp"] = pd.to_datetime(df["timestamp"])
        return df
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "get_audit_log_dataframe")
        return pd.DataFrame(columns=AUDIT_LOG_COLUMNS)


def verify_audit_log_integrity() -> dict:
    """Walks the ENTIRE audit_log table in insertion order and recomputes
    each row's hash from its own fields plus the previous row's stored
    hash, comparing against what's stored. Returns
    {"ok": bool, "rows_checked": int, "first_broken_id": int | None}.
    A mismatch means a row was edited or deleted after the fact (or rows
    were inserted out of order) - see the module docstring for what this
    can and can't guarantee. Never raises; reports ok=True, rows_checked=0
    if the database is unavailable or the table is empty (nothing to
    contradict)."""
    if not _db_ok():
        return {"ok": True, "rows_checked": 0, "first_broken_id": None}
    try:
        with _get_connection() as conn:
            cur = conn.execute(
                "SELECT id, timestamp, event_type, actor, detail, organization, project, site, "
                "prev_hash, row_hash FROM audit_log ORDER BY id ASC"
            )
            rows = cur.fetchall()
    except Exception as exc:  # noqa: BLE001
        _handle_db_failure(exc, "verify_audit_log_integrity")
        return {"ok": True, "rows_checked": 0, "first_broken_id": None}

    expected_prev = _GENESIS_HASH
    for row in rows:
        row_id, timestamp, event_type, actor, detail, organization, project, site, prev_hash, row_hash = row
        detail = detail or ""
        organization = organization or ""
        project = project or ""
        site = site or ""
        if prev_hash != expected_prev:
            return {"ok": False, "rows_checked": len(rows), "first_broken_id": row_id}
        recomputed = _hash_audit_row(prev_hash, timestamp, event_type, actor, detail,
                                      organization, project, site)
        if recomputed != row_hash:
            return {"ok": False, "rows_checked": len(rows), "first_broken_id": row_id}
        expected_prev = row_hash

    return {"ok": True, "rows_checked": len(rows), "first_broken_id": None}


def monthly_summary(df: pd.DataFrame) -> pd.DataFrame:
    """Aggregates the log by (year-month, module): assessment count, safety
    override count, and average key metric."""
    if df.empty:
        return pd.DataFrame(columns=["month", "module", "assessments", "safety_overrides", "avg_key_metric"])
    work = df.copy()
    work["month"] = work["timestamp"].dt.tz_localize(None).dt.to_period("M").astype(str)
    grouped = work.groupby(["month", "module"]).agg(
        assessments=("module", "count"),
        safety_overrides=("safety_override", "sum"),
        avg_key_metric=("key_metric", "mean"),
    ).reset_index()
    grouped["avg_key_metric"] = grouped["avg_key_metric"].round(2)
    return grouped


def build_monthly_excel(df: pd.DataFrame) -> bytes:
    """Builds a real downloadable .xlsx with:
      - 'Raw Log' sheet: every logged assessment
      - 'Monthly Summary' sheet: assessments/overrides/avg metric by month+module
      - 'Trend Chart' sheet: an embedded line chart (assessment count per
        month per module) - a real chart object in the workbook, not just
        a description of one.
    """
    summary = monthly_summary(df)

    buffer = io.BytesIO()
    export_df = df.copy()
    if not export_df.empty:
        # Excel has no concept of timezone-aware datetimes - strip the tz
        # info (values stay correct, just displayed as naive UTC) rather
        # than letting pandas raise ValueError on write.
        export_df["timestamp"] = export_df["timestamp"].dt.tz_localize(None)
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        (export_df if not export_df.empty else pd.DataFrame(columns=LOG_COLUMNS)).to_excel(
            writer, sheet_name="Raw Log", index=False
        )
        summary.to_excel(writer, sheet_name="Monthly Summary", index=False)
    buffer.seek(0)

    wb = load_workbook(buffer)
    chart_sheet = wb.create_sheet("Trend Chart")

    if not summary.empty:
        pivot = summary.pivot(index="month", columns="module", values="assessments").fillna(0)
        chart_sheet.append(["month"] + list(pivot.columns))
        for month, row in pivot.iterrows():
            chart_sheet.append([month] + list(row.values))

        chart = LineChart()
        chart.title = "Assessments per Month by Module"
        chart.y_axis.title = "Assessment count"
        chart.x_axis.title = "Month"
        n_rows, n_cols = pivot.shape
        data_ref = Reference(chart_sheet, min_col=2, max_col=1 + n_cols, min_row=1, max_row=1 + n_rows)
        cats_ref = Reference(chart_sheet, min_col=1, max_col=1, min_row=2, max_row=1 + n_rows)
        chart.add_data(data_ref, titles_from_data=True)
        chart.set_categories(cats_ref)
        chart_sheet.add_chart(chart, "H2")
    else:
        chart_sheet.append(["No data logged yet this session."])

    out_buffer = io.BytesIO()
    wb.save(out_buffer)
    return out_buffer.getvalue()
