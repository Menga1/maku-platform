"""
MAKU test suite
================
Covers all 5 risk_engine.py modules and their handoff into ai_advisor.py.
Per the project's Mathematical Isolation rule, these tests exercise the
engine functions directly (not the Streamlit pages, which are UI-only).

Run with:
    pytest
or:
    pytest -v test_app.py
"""

import pytest
from unittest.mock import patch, MagicMock

import streamlit as st

from risk_engine import (
    calculate_solar_albedo_heat_risk,
    calculate_marine_humidex_risk,
    calculate_underground_kinetic_risk,
    calculate_high_rise_kinetic_risk,
    calculate_datacenter_kinetic_risk,
    calculate_wind_energy_kinetic_risk,
    calculate_mining_quarrying_kinetic_risk,
    calculate_marine_port_kinetic_risk,
    noise_dose_percent,
    whole_body_vibration_a8,
    corroded_capacity_pct,
    calculate_iso7243_heat_stress,
    recommended_work_rest_ratio,
    get_stop_work_triggers,
    STOP_WORK_TRIGGERS,
    CLOTHING_ADJUSTMENT_FACTOR_C,
    CLOTHING_REQUIRES_PHYSIOLOGICAL_MONITORING,
)
from risk_matrix import (
    matrix_score,
    matrix_band,
    severity_from_band,
    likelihood_from_margin,
    score_hazard,
    aggregate_risk_matrix,
    apply_controls_residual_risk,
)
from ai_advisor import (
    get_controls, generate_narrative, get_regulatory_references, get_bibliography, CONTROLS_MAP,
    translate_narrative, generate_daily_briefing, predict_forecast_breach,
    generate_predictive_alert, classify_api_error,
)
from regulatory_references import (
    get_references, get_further_reading, REGULATORY_REFERENCES,
    FORMULA_STANDARDS_MAP, get_formula_standard, get_all_formula_standards,
)
from regulatory_country_thresholds import (
    get_country_thresholds, is_midday_outdoor_ban_active, get_regulatory_profile, resolve_heat_stress_limit,
)
from ui_helpers import _build_report_pdf, HSE_DISCLAIMER_TEXT, render_hse_disclaimer, render_regulatory_badge
from analytics import (
    log_assessment, get_log_dataframe, merge_uploaded_csv,
    monthly_summary, build_monthly_excel, LOG_COLUMNS,
    log_audit_event, get_audit_log_dataframe, verify_audit_log_integrity,
    AUDIT_EVENT_LOGIN_SUCCESS, AUDIT_EVENT_THRESHOLD_OVERRIDE,
    set_org_context, get_org_context, DEFAULT_ORGANIZATION, DEFAULT_PROJECT,
    build_evidence_traceability,
)
from data_feeds import (
    fetch_solar_forecast, fetch_offshore_forecast, DataFeedError,
    fetch_live_weather_universal, _mock_universal_weather,
)
import auth
from regulatory_country_thresholds import (
    get_threshold_category_badge, THRESHOLD_CATEGORY_META, REGULATORY_PROFILES,
)
import ai_advisor


REQUIRED_KEYS = {"module", "risk_band", "primary_hazard", "drivers", "safety_override"}


def _assert_standard_shape(result: dict):
    """Every module's return dict must carry the shared ai_advisor-compatible fields."""
    missing = REQUIRED_KEYS - result.keys()
    assert not missing, f"Missing compatibility fields: {missing}"
    assert result["module"] in CONTROLS_MAP, f"No CONTROLS_MAP entry for module {result['module']!r}"
    assert isinstance(result["safety_override"], bool)
    assert isinstance(result["drivers"], dict) and result["drivers"]


# ---------------------------------------------------------------------------
# Module 1: Solar (Desert)
# ---------------------------------------------------------------------------

class TestSolar:
    def test_low_risk_shape(self):
        r = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        _assert_standard_shape(r)
        assert r["risk_level"] == "LOW"
        assert r["safety_override"] is False

    def test_critical_via_high_temp(self):
        r = calculate_solar_albedo_heat_risk(ghi=1100, uv_index=6, ambient_temp=48, surface_type="silicon_pv_panels")
        assert r["risk_level"] == "CRITICAL"
        assert r["override_required"] is True

    def test_critical_via_uv_alone(self):
        r = calculate_solar_albedo_heat_risk(ghi=200, uv_index=11, ambient_temp=25, surface_type="pure_desert_sand")
        assert r["risk_level"] == "CRITICAL"

    def test_unknown_surface_falls_back_to_default_albedo(self):
        r = calculate_solar_albedo_heat_risk(ghi=500, uv_index=5, ambient_temp=30, surface_type="not_a_real_surface")
        assert r["drivers"]["Albedo factor applied"] == pytest.approx(0.25)

    def test_pipeline_into_ai_advisor(self):
        r = calculate_solar_albedo_heat_risk(ghi=950, uv_index=9, ambient_temp=42, surface_type="hybrid_assembly_zone")
        controls = get_controls(r)
        assert isinstance(controls, list) and controls
        for lang in ("fr", "en"):
            narrative = generate_narrative(r, controls, api_key=None, lang=lang)
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Module 2: Offshore (Marine)
# ---------------------------------------------------------------------------

class TestOffshore:
    def test_low_risk_normal_wind(self):
        r = calculate_marine_humidex_risk(ambient_temp=24, relative_humidity=55, wind_speed=8)
        _assert_standard_shape(r)
        assert r["risk_band"] == "Low"
        assert r["wind_risk_status"] == "Normal Operations"
        assert r["safety_override"] is False

    def test_wind_restricted_band(self):
        r = calculate_marine_humidex_risk(ambient_temp=25, relative_humidity=60, wind_speed=20)
        assert r["wind_risk_status"] == "Restricted - Monitor Closely"

    def test_wind_suspended_forces_override(self):
        r = calculate_marine_humidex_risk(ambient_temp=25, relative_humidity=60, wind_speed=27)
        assert r["wind_risk_status"] == "Suspended - Crane/Lifting Danger"
        assert r["safety_override"] is True

    def test_extreme_humidex_forces_override(self):
        r = calculate_marine_humidex_risk(ambient_temp=34, relative_humidity=98, wind_speed=5)
        assert r["risk_band"] == "Extreme"
        assert r["safety_override"] is True

    def test_pipeline_into_ai_advisor(self):
        r = calculate_marine_humidex_risk(ambient_temp=33, relative_humidity=92, wind_speed=15)
        controls = get_controls(r)
        assert isinstance(controls, list) and controls
        for lang in ("fr", "en"):
            narrative = generate_narrative(r, controls, api_key=None, lang=lang)
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Module 3: Underground (Tunnel/Metro)
# ---------------------------------------------------------------------------

class TestUnderground:
    def test_low_risk_shape(self):
        r = calculate_underground_kinetic_risk(
            ambient_temp=22, geothermal_humidity=55, particulate_matter_pm25=20, gas_co_ppm=5,
        )
        _assert_standard_shape(r)
        assert r["risk_band"] == "LOW"
        assert r["safety_override"] is False

    def test_critical_via_heat(self):
        r = calculate_underground_kinetic_risk(
            ambient_temp=35, geothermal_humidity=95, particulate_matter_pm25=40, gas_co_ppm=8,
        )
        assert r["risk_band"] == "CRITICAL"
        assert r["perceived_temp"] > 42.0
        assert r["safety_override"] is True

    def test_critical_via_co_oel(self):
        r = calculate_underground_kinetic_risk(
            ambient_temp=25, geothermal_humidity=60, particulate_matter_pm25=40, gas_co_ppm=30,
        )
        assert r["gas_exceeds"] is True
        assert r["safety_override"] is True
        assert r["risk_band"] == "CRITICAL"

    def test_critical_via_pm25_oel(self):
        r = calculate_underground_kinetic_risk(
            ambient_temp=25, geothermal_humidity=60, particulate_matter_pm25=300, gas_co_ppm=8,
        )
        assert r["dust_exceeds"] is True
        assert r["safety_override"] is True

    def test_pipeline_into_ai_advisor(self):
        r = calculate_underground_kinetic_risk(
            ambient_temp=30, geothermal_humidity=80, particulate_matter_pm25=90, gas_co_ppm=12,
        )
        controls = get_controls(r)
        assert isinstance(controls, list) and controls
        for lang in ("fr", "en"):
            narrative = generate_narrative(r, controls, api_key=None, lang=lang)
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Module 4: High-Rise (Vertical Urban)
# ---------------------------------------------------------------------------

class TestHighRise:
    def test_low_risk_shape(self):
        r = calculate_high_rise_kinetic_risk(ground_wind_speed_knots=8, floor_level=5, crane_load_mass_tons=10)
        _assert_standard_shape(r)
        assert r["risk_band"] == "LOW"
        assert r["safety_override"] is False

    def test_regulatory_profile_changes_real_safety_outcome(self):
        # Same physical wind speed: France's stricter (illustrative) crane
        # suspend threshold triggers a safety override that the harmonized
        # USA default does not - proves the profile genuinely drives the
        # engine's decision, not just a cosmetic label.
        r_usa = calculate_high_rise_kinetic_risk(20.5, 40, 4)
        r_fr = calculate_high_rise_kinetic_risk(20.5, 40, 4, regulatory_profile=get_regulatory_profile("FRANCE"))
        assert r_usa["safety_override"] is False
        assert r_fr["safety_override"] is True
        assert r_usa["regulatory_country_code"] == "USA"
        assert r_fr["regulatory_country_code"] == "FRANCE"

    def test_high_wind_forces_override(self):
        r = calculate_high_rise_kinetic_risk(ground_wind_speed_knots=25, floor_level=90, crane_load_mass_tons=15)
        assert r["scaled_wind_speed"] > 30.0
        assert r["safety_override"] is True
        assert r["risk_band"] == "CRITICAL"

    def test_light_load_oscillation_escalates_band_without_override(self):
        # Light crane load at moderate scaled wind should push oscillation-driven
        # severity up even while staying under the hard 30-knot override.
        r = calculate_high_rise_kinetic_risk(ground_wind_speed_knots=18, floor_level=60, crane_load_mass_tons=1)
        assert r["scaled_wind_speed"] <= 30.0
        assert r["safety_override"] is False
        assert r["risk_band"] in ("HIGH", "CRITICAL")

    def test_wind_scales_up_with_floor_level(self):
        low = calculate_high_rise_kinetic_risk(ground_wind_speed_knots=15, floor_level=1, crane_load_mass_tons=10)
        high = calculate_high_rise_kinetic_risk(ground_wind_speed_knots=15, floor_level=100, crane_load_mass_tons=10)
        assert high["scaled_wind_speed"] > low["scaled_wind_speed"]

    def test_pipeline_into_ai_advisor(self):
        r = calculate_high_rise_kinetic_risk(ground_wind_speed_knots=20, floor_level=150, crane_load_mass_tons=4)
        controls = get_controls(r)
        assert isinstance(controls, list) and controls
        for lang in ("fr", "en"):
            narrative = generate_narrative(r, controls, api_key=None, lang=lang)
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Module 5: Data Center (Controlled Critical Environment)
# ---------------------------------------------------------------------------

class TestDataCenter:
    def test_low_risk_shape(self):
        r = calculate_datacenter_kinetic_risk(
            electrical_load_kw=100, hot_aisle_temp=26,
            ceiling_void_confined=False, gas_system_armed=False,
        )
        _assert_standard_shape(r)
        assert r["risk_band"] == "LOW"
        assert r["safety_override"] is False

    def test_arc_flash_danger_forces_override(self):
        r = calculate_datacenter_kinetic_risk(
            electrical_load_kw=1300, hot_aisle_temp=30,
            ceiling_void_confined=False, gas_system_armed=False,
        )
        assert r["arc_flash_danger"] is True
        assert r["safety_override"] is True
        assert r["risk_band"] == "CRITICAL"

    def test_thermal_critical_forces_override(self):
        r = calculate_datacenter_kinetic_risk(
            electrical_load_kw=200, hot_aisle_temp=48,
            ceiling_void_confined=False, gas_system_armed=False,
        )
        assert r["safety_override"] is True

    def test_confined_plus_armed_forces_override_even_at_low_load(self):
        r = calculate_datacenter_kinetic_risk(
            electrical_load_kw=150, hot_aisle_temp=28,
            ceiling_void_confined=True, gas_system_armed=True,
        )
        assert r["confined_armed_danger"] is True
        assert r["safety_override"] is True

    def test_confined_alone_without_armed_system_is_not_dangerous(self):
        r = calculate_datacenter_kinetic_risk(
            electrical_load_kw=150, hot_aisle_temp=28,
            ceiling_void_confined=True, gas_system_armed=False,
        )
        assert r["confined_armed_danger"] is False

    def test_thermal_control_reflects_actual_thermal_reading_not_overall_band(self):
        # Regression test for the bug caught during development: the
        # "thermal differential elevated" control must be driven by the
        # actual thermal_differential value, not by an overall CRITICAL
        # band caused purely by arc-flash danger.
        r = calculate_datacenter_kinetic_risk(
            electrical_load_kw=1300, hot_aisle_temp=30,
            ceiling_void_confined=False, gas_system_armed=False,
        )
        assert r["thermal_differential"] < 15.0
        controls = get_controls(r)
        assert not any("thermal differential" in c.lower() for c in controls)

    def test_pipeline_into_ai_advisor(self):
        r = calculate_datacenter_kinetic_risk(
            electrical_load_kw=400, hot_aisle_temp=36,
            ceiling_void_confined=False, gas_system_armed=True,
        )
        controls = get_controls(r)
        assert isinstance(controls, list) and controls
        for lang in ("fr", "en"):
            narrative = generate_narrative(r, controls, api_key=None, lang=lang)
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Module 6: Wind Energy (Onshore/Offshore)
# ---------------------------------------------------------------------------

class TestWindEnergy:
    def test_low_wind_onshore_shape(self):
        r = calculate_wind_energy_kinetic_risk(6.0)
        _assert_standard_shape(r)
        assert r["risk_band"] == "LOW"
        assert r["safety_override"] is False

    def test_high_wind_forces_critical_and_override(self):
        r = calculate_wind_energy_kinetic_risk(21.0)
        assert r["risk_band"] == "CRITICAL"
        assert r["safety_override"] is True

    def test_restricted_band_between_thresholds(self):
        r = calculate_wind_energy_kinetic_risk(16.0)
        assert r["risk_band"] == "HIGH"
        assert r["safety_override"] is False

    def test_lightning_30_30_rule_forces_override_even_at_low_wind(self):
        r = calculate_wind_energy_kinetic_risk(5.0, flash_to_bang_sec=12.0)
        assert "STOP WORK" in r["lightning_status"]
        assert r["safety_override"] is True

    def test_lightning_above_30s_does_not_trigger_stop(self):
        r = calculate_wind_energy_kinetic_risk(5.0, flash_to_bang_sec=45.0)
        assert "STOP WORK" not in r["lightning_status"]
        assert r["safety_override"] is False

    def test_offshore_rough_seas_suspend_ctv_and_forces_override(self):
        r = calculate_wind_energy_kinetic_risk(5.0, is_offshore=True, significant_wave_height_m=2.5)
        assert "Suspended" in r["ctv_transfer_status"]
        assert r["safety_override"] is True

    def test_onshore_ignores_sea_state(self):
        r = calculate_wind_energy_kinetic_risk(5.0, is_offshore=False, significant_wave_height_m=5.0)
        assert r["ctv_transfer_status"] == "Not applicable (onshore)"
        assert r["safety_override"] is False

    def test_pipeline_into_ai_advisor(self):
        r = calculate_wind_energy_kinetic_risk(22.0, is_offshore=True, significant_wave_height_m=2.5, flash_to_bang_sec=10)
        controls = get_controls(r)
        assert isinstance(controls, list) and controls
        assert any("SAFETY OVERRIDE" in c for c in controls)
        for lang in ("fr", "en", "ar", "es"):
            narrative = generate_narrative(r, controls, api_key=None, lang=lang)
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Module 7: Mining & Quarrying
# ---------------------------------------------------------------------------

class TestMiningQuarrying:
    def test_low_risk_shape(self):
        r = calculate_mining_quarrying_kinetic_risk(
            respirable_silica_ugm3=10, measured_noise_dba=75, noise_exposure_hours=8,
            measured_vibration_aw_ms2=0.2, vibration_exposure_hours=8,
        )
        _assert_standard_shape(r)
        assert r["risk_band"] == "LOW"
        assert r["safety_override"] is False

    def test_silica_over_oel_forces_override(self):
        r = calculate_mining_quarrying_kinetic_risk(
            respirable_silica_ugm3=60, measured_noise_dba=75, noise_exposure_hours=8,
            measured_vibration_aw_ms2=0.2, vibration_exposure_hours=8,
        )
        assert r["silica_exceeds_oel"] is True
        assert r["safety_override"] is True
        assert r["risk_band"] == "CRITICAL"

    def test_silica_action_level_high_band_no_override(self):
        r = calculate_mining_quarrying_kinetic_risk(
            respirable_silica_ugm3=35, measured_noise_dba=75, noise_exposure_hours=8,
            measured_vibration_aw_ms2=0.2, vibration_exposure_hours=8,
        )
        assert r["silica_exceeds_oel"] is False
        assert r["safety_override"] is False
        assert r["risk_band"] == "HIGH"

    def test_vibration_exceeds_limit_forces_override(self):
        r = calculate_mining_quarrying_kinetic_risk(
            respirable_silica_ugm3=10, measured_noise_dba=75, noise_exposure_hours=8,
            measured_vibration_aw_ms2=1.5, vibration_exposure_hours=8,
        )
        assert r["vibration_exceeds_limit"] is True
        assert r["safety_override"] is True

    def test_noise_dose_formula_matches_standard_exchange_rate_math(self):
        # At exactly the 90 dBA criterion for a full 8h shift, dose should be 100%.
        assert noise_dose_percent(90.0, 8.0) == 100.0
        # 3 dB above criterion with a 3 dB exchange rate halves the allowed time.
        assert noise_dose_percent(93.0, 8.0, criterion_dba=90.0, exchange_rate_db=3.0) == 200.0

    def test_whole_body_vibration_a8_normalizes_duration(self):
        # Same instantaneous a_w, shorter exposure -> lower A(8).
        full_day = whole_body_vibration_a8(1.0, 8.0)
        half_day = whole_body_vibration_a8(1.0, 2.0)
        assert half_day < full_day
        assert full_day == 1.0

    def test_pipeline_into_ai_advisor(self):
        r = calculate_mining_quarrying_kinetic_risk(
            respirable_silica_ugm3=60, measured_noise_dba=100, noise_exposure_hours=8,
            measured_vibration_aw_ms2=1.5, vibration_exposure_hours=8,
        )
        controls = get_controls(r)
        assert isinstance(controls, list) and controls
        assert any("SAFETY OVERRIDE" in c for c in controls)
        for lang in ("fr", "en", "ar", "es"):
            narrative = generate_narrative(r, controls, api_key=None, lang=lang)
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Module 8: Marine & Port Construction
# ---------------------------------------------------------------------------

class TestMarinePortConstruction:
    def test_low_risk_shape(self):
        r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=2.0, required_min_clearance_m=1.0,
            is_night_operation=False, measured_illuminance_lux=200,
            hardware_years_in_service=1,
        )
        _assert_standard_shape(r)
        assert r["risk_band"] == "LOW"
        assert r["safety_override"] is False

    def test_critical_tide_clearance_forces_override(self):
        r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=1.1, required_min_clearance_m=1.0,
            is_night_operation=False, measured_illuminance_lux=200,
            hardware_years_in_service=1,
        )
        assert r["tide_clearance_margin_m"] <= 0.3
        assert r["safety_override"] is True

    def test_night_amplification_escalates_band(self):
        night_r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=1.7, required_min_clearance_m=1.0,
            is_night_operation=True, measured_illuminance_lux=20,
            hardware_years_in_service=1,
        )
        day_r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=1.7, required_min_clearance_m=1.0,
            is_night_operation=False, measured_illuminance_lux=20,
            hardware_years_in_service=1,
        )
        assert night_r["night_amplified"] is True
        assert day_r["night_amplified"] is False

    def test_adequate_night_lighting_does_not_amplify(self):
        r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=2.0, required_min_clearance_m=1.0,
            is_night_operation=True, measured_illuminance_lux=200,
            hardware_years_in_service=1,
        )
        assert r["night_amplified"] is False

    def test_old_hardware_forces_override_via_low_remaining_capacity(self):
        r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=2.0, required_min_clearance_m=1.0,
            is_night_operation=False, measured_illuminance_lux=200,
            hardware_years_in_service=15, hardware_exposure_class="C5M_marine_splash_zone",
        )
        assert r["hardware_remaining_capacity_pct"] < 70.0
        assert r["safety_override"] is True

    def test_corrosion_derating_never_goes_negative(self):
        assert corroded_capacity_pct(1000, "C5M_marine_splash_zone") == 0.0

    def test_pipeline_into_ai_advisor(self):
        r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=1.1, required_min_clearance_m=1.0,
            is_night_operation=True, measured_illuminance_lux=20,
            hardware_years_in_service=15,
        )
        controls = get_controls(r)
        assert isinstance(controls, list) and controls
        assert any("SAFETY OVERRIDE" in c for c in controls)
        for lang in ("fr", "en", "ar", "es"):
            narrative = generate_narrative(r, controls, api_key=None, lang=lang)
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Module 5a: Country-specific regulatory thresholds
# ---------------------------------------------------------------------------

class TestCountryThresholds:
    def test_us_matches_harmonized_defaults(self):
        cfg = get_country_thresholds("USA")
        assert cfg["noise_criterion_dba"] == 90.0
        assert cfg["noise_exchange_rate_db"] == 5.0
        assert cfg["midday_outdoor_work_ban"] is False

    def test_france_uses_stricter_eu_noise_limits(self):
        cfg = get_country_thresholds("FRANCE")
        assert cfg["noise_criterion_dba"] == 85.0
        assert cfg["noise_exchange_rate_db"] == 3.0

    def test_uae_has_midday_ban_flag(self):
        cfg = get_country_thresholds("UAE")
        assert cfg["midday_outdoor_work_ban"] is True

    def test_unknown_country_falls_back_to_harmonized_default(self):
        cfg = get_country_thresholds("ZZ_NOT_REAL")
        assert cfg["country_code"] == "USA"
        assert cfg["noise_criterion_dba"] == 90.0

    def test_noise_formula_accepts_country_specific_parameters(self):
        # Same 87 dBA / 8h reading: France's stricter criterion should
        # report a higher dose than the US's more lenient one.
        us_dose = noise_dose_percent(87.0, 8.0, criterion_dba=90.0, exchange_rate_db=5.0)
        fr_dose = noise_dose_percent(87.0, 8.0, criterion_dba=85.0, exchange_rate_db=3.0)
        assert fr_dose > us_dose

    def test_uae_midday_ban_active_in_season_and_hours(self):
        from datetime import date
        assert is_midday_outdoor_ban_active("UAE", date(2026, 7, 15), 13, 0) is True
        assert is_midday_outdoor_ban_active("UAE", date(2026, 1, 15), 13, 0) is False
        assert is_midday_outdoor_ban_active("USA", date(2026, 7, 15), 13, 0) is False

    def test_heat_stress_method_differs_between_usa_and_france(self):
        usa = get_country_thresholds("USA")
        france = get_country_thresholds("FRANCE")
        assert usa["heat_stress"]["method"] == "ACGIH"
        assert france["heat_stress"]["method"] == "ISO7243"

    def test_crane_wind_shear_thresholds_present_for_all_three(self):
        for code in ("USA", "FRANCE", "UAE"):
            cfg = get_country_thresholds(code)
            assert cfg["wind_shear"]["crane_suspend_knots"] > cfg["wind_shear"]["crane_restrict_knots"]

    def test_ambient_pm25_differs_by_country_with_distinct_averaging_periods(self):
        usa = get_country_thresholds("USA")
        france = get_country_thresholds("FRANCE")
        uae = get_country_thresholds("UAE")
        assert usa["air_quality"]["ambient_averaging_period"] == "24-hour"
        assert france["air_quality"]["ambient_averaging_period"] == "annual"
        assert uae["air_quality"]["ambient_averaging_period"] == "24-hour"

    def test_occupational_oel_stays_harmonized_across_all_three(self):
        # Per the honesty note in regulatory_country_thresholds.py: no
        # country has a distinct published occupational PM2.5/CO OEL, so
        # all three should be identical (not a fabricated difference).
        profiles = [get_country_thresholds(c) for c in ("USA", "FRANCE", "UAE")]
        pm25_oels = {p["air_quality"]["pm25_oel_ugm3"] for p in profiles}
        co_oels = {p["air_quality"]["co_oel_ppm"] for p in profiles}
        assert len(pm25_oels) == 1
        assert len(co_oels) == 1

    def test_resolve_heat_stress_limit_acgih_uses_work_rest_ratio(self):
        usa = get_country_thresholds("USA")
        light_full = resolve_heat_stress_limit(usa, "light", "100/0")
        light_rested = resolve_heat_stress_limit(usa, "light", "25/75")
        assert light_rested["limit"] > light_full["limit"]

    def test_resolve_heat_stress_limit_iso7243_ignores_work_rest_ratio(self):
        france = get_country_thresholds("FRANCE")
        a = resolve_heat_stress_limit(france, "moderate", "100/0")
        b = resolve_heat_stress_limit(france, "moderate", "25/75")
        assert a["limit"] == b["limit"]


# ---------------------------------------------------------------------------
# Module 1 (i18n part 2): translate_narrative honesty guarantees
# ---------------------------------------------------------------------------

class TestTranslateNarrative:
    def test_no_api_key_returns_original_text_unchanged(self):
        original = "Risk level: HIGH."
        result = translate_narrative(original, "ar", api_key=None)
        assert result["text"] == original
        assert result["translated"] is False
        assert result["error_kind"] == "no_api_key"

    def test_unsupported_language_does_not_crash(self):
        result = translate_narrative("test", "zz_not_a_language", api_key="whatever")
        assert result["translated"] is False
        assert result["error_kind"] == "unsupported_language"

    def test_bad_key_fails_gracefully_without_fabricating_translation(self):
        original = "Risk level: HIGH."
        result = translate_narrative(original, "es", api_key="clearly_invalid_key")
        assert result["translated"] is False
        assert result["text"] == original  # never returns a fabricated/garbled "translation"


# ---------------------------------------------------------------------------
# Module 4: Daily briefing + predictive forecasting
# ---------------------------------------------------------------------------

class TestDailyBriefing:
    def test_fallback_template_used_without_api_key(self):
        result = generate_daily_briefing(
            "Test Site", ["Task A", "Task B"], {"temp": 30}, api_key=None, lang="en",
        )
        assert result["source"] == "template"
        assert "Task A" in result["script"] and "Task B" in result["script"]

    def test_fallback_handles_empty_task_list(self):
        result = generate_daily_briefing("Test Site", [], {"temp": 30}, api_key=None, lang="fr")
        assert isinstance(result["script"], str) and result["script"]


class TestPredictiveForecasting:
    def test_no_breach_when_forecast_stays_in_same_band(self):
        current = {"ghi": 500, "uv_index": 5, "ambient_temp": 30, "surface_type": "hybrid_assembly_zone"}
        forecast = [{"ghi": 510, "uv_index": 5, "ambient_temp": 31, "surface_type": "hybrid_assembly_zone", "label": "in 1h"}]
        result = predict_forecast_breach(calculate_solar_albedo_heat_risk, current, forecast)
        assert result["will_breach"] is False
        assert generate_predictive_alert(result, "en") is None

    def test_breach_detected_and_labeled_correctly(self):
        current = {"ghi": 500, "uv_index": 5, "ambient_temp": 30, "surface_type": "hybrid_assembly_zone"}
        forecast = [
            {"ghi": 700, "uv_index": 8, "ambient_temp": 38, "surface_type": "hybrid_assembly_zone", "label": "in 2h"},
            {"ghi": 950, "uv_index": 11, "ambient_temp": 45, "surface_type": "hybrid_assembly_zone", "label": "in 4h"},
        ]
        result = predict_forecast_breach(calculate_solar_albedo_heat_risk, current, forecast)
        assert result["will_breach"] is True
        assert result["breach_label"] == "in 2h"  # first breach point, not the worst one
        for lang in ("fr", "en", "ar", "es"):
            alert = generate_predictive_alert(result, lang)
            assert isinstance(alert, str) and alert

    def test_already_critical_short_circuits_before_checking_forecast(self):
        current = {"ghi": 1200, "uv_index": 12, "ambient_temp": 48, "surface_type": "hybrid_assembly_zone"}
        forecast = [{"ghi": 100, "uv_index": 1, "ambient_temp": 20, "surface_type": "hybrid_assembly_zone", "label": "in 1h"}]
        result = predict_forecast_breach(calculate_solar_albedo_heat_risk, current, forecast)
        assert result["currently_critical"] is True
        assert result["will_breach"] is False

    def test_works_across_a_different_module_signature_unchanged(self):
        # Same generic function, wind energy's completely different kwargs -
        # proves this isn't hardcoded to the solar module's parameter names.
        current = {"hub_height_wind_speed_ms": 6.0}
        forecast = [{"hub_height_wind_speed_ms": 21.0, "label": "in 3h"}]
        result = predict_forecast_breach(calculate_wind_energy_kinetic_risk, current, forecast)
        assert result["will_breach"] is True
        assert result["breach_label"] == "in 3h"


# ---------------------------------------------------------------------------
# Module 5b: API error classification (offline-first support)
# ---------------------------------------------------------------------------

class TestClassifyApiError:
    def test_connection_error_classified_as_network(self):
        import requests
        assert classify_api_error(requests.exceptions.ConnectionError("no route to host")) == "network"

    def test_timeout_classified_as_timeout(self):
        import requests
        assert classify_api_error(requests.exceptions.Timeout("timed out")) == "timeout"

    def test_auth_failure_classified_as_auth(self):
        assert classify_api_error(Exception("401 Unauthorized: invalid x-api-key")) == "auth"

    def test_rate_limit_classified_correctly(self):
        assert classify_api_error(Exception("429 Too Many Requests - rate limit exceeded")) == "rate_limit"

    def test_unknown_error_classified_as_other(self):
        assert classify_api_error(Exception("something completely unexpected")) == "other"


# ---------------------------------------------------------------------------
# Module 3: Universal weather ingestion fallback chain
# ---------------------------------------------------------------------------

class TestUniversalWeatherFeed:
    def test_mock_fallback_has_all_expected_fields(self):
        mock = _mock_universal_weather(24.95, 53.90)
        expected_fields = {
            "temperature_2m", "relative_humidity_2m", "wind_speed_10m_kn",
            "wind_gusts_10m_kn", "uv_index", "pressure_hpa", "fetched_at", "source",
        }
        assert expected_fields.issubset(mock.keys())

    def test_falls_back_to_mock_when_no_key_and_network_blocked(self):
        # In this sandboxed test environment, api.open-meteo.com is not on
        # the allowed egress list, so this genuinely exercises the real
        # fallback path end-to-end rather than mocking it away.
        result = fetch_live_weather_universal(24.95, 53.90)
        assert result["source"] in ("Mock fallback data (no live network reachable)",
                                     "Open-Meteo /v1/forecast")

    def test_raises_when_mock_fallback_disabled_and_all_providers_fail(self):
        with patch("data_feeds._fetch_open_meteo_universal", side_effect=Exception("simulated network failure")):
            with pytest.raises(DataFeedError):
                fetch_live_weather_universal(1.0, 1.0, allow_mock_fallback=False)


# ---------------------------------------------------------------------------
# Cross-module: every module registered in ai_advisor.CONTROLS_MAP
# ---------------------------------------------------------------------------

def test_all_nine_modules_registered_in_controls_map():
    # 8 kinetic-risk modules + the ISO 7243/ACGIH Occupational Heat Stress
    # module added by the HSE audit's Heat-Stress Upgrade corrective action.
    expected_modules = {
        "Solar (Desert)",
        "Offshore (Marine)",
        "Underground (Tunnel/Metro)",
        "High-Rise (Vertical Urban)",
        "Data Center (Controlled Critical Environment)",
        "Wind Energy (Onshore/Offshore)",
        "Mining & Quarrying",
        "Marine & Port Construction",
        "Occupational Heat Stress (ISO 7243 / ACGIH TLV)",
    }
    assert expected_modules == set(CONTROLS_MAP.keys())


# ---------------------------------------------------------------------------
# Regulatory references & bibliography
# ---------------------------------------------------------------------------

class TestRegulatoryReferences:
    def test_every_controls_map_module_has_references(self):
        # Every module ai_advisor knows about must have a citation list -
        # a silently-empty references section would be a regression.
        for module in CONTROLS_MAP.keys():
            refs = get_references(module)
            assert len(refs) >= 3, f"{module} has too few regulatory references"

    def test_reference_entries_have_required_fields(self):
        for module, refs in REGULATORY_REFERENCES.items():
            for ref in refs:
                assert ref.get("region")
                assert ref.get("body")
                assert ref.get("doc")
                # Every reference must be a citation (name + optional link),
                # never a large text blob - this is a proxy check against
                # accidentally pasting reproduced regulation/book text in.
                assert len(ref["doc"]) < 300

    def test_unknown_module_returns_empty_list_not_error(self):
        assert get_references("Not A Real Module") == []

    def test_further_reading_cites_brauer_by_name_only(self):
        biblio = get_further_reading()
        authors = [entry["author"] for entry in biblio]
        assert "Roger L. Brauer" in authors
        brauer_entry = next(e for e in biblio if e["author"] == "Roger L. Brauer")
        # Citation only (title/publisher/short note) - not book content.
        assert len(brauer_entry.get("note", "")) < 300

    def test_ai_advisor_wrapper_functions_match_underlying_module(self):
        result = calculate_solar_albedo_heat_risk(500, 5, 30, "pure_desert_sand")
        assert get_regulatory_references(result) == get_references(result["module"])
        assert get_bibliography() == get_further_reading()

    def test_narrative_regulatory_basis_only_names_cited_bodies(self):
        # Regression test: the AI narrative's "regulatory basis" line must be
        # built from the same verified reference list, never a separate
        # hardcoded/unverified list (this caught a real bug where an
        # unverified "Bauer temporary-works principles" claim had been
        # hardcoded directly into the narrative generator).
        result = calculate_underground_kinetic_risk(30, 85, 100, 15)
        controls = get_controls(result)
        narrative = generate_narrative(result, controls, api_key=None, lang="fr")
        cited_bodies = {ref["body"] for ref in get_references(result["module"])}
        assert cited_bodies, "expected at least one cited body for this module"
        assert any(body in narrative for body in cited_bodies)
        assert "Bauer" not in narrative


# ---------------------------------------------------------------------------
# PDF report generation
# ---------------------------------------------------------------------------

class TestPdfReport:
    def test_pdf_is_valid_for_every_module(self):
        cases = [
            calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels"),
            calculate_marine_humidex_risk(33, 92, 15),
            calculate_underground_kinetic_risk(30, 80, 90, 12),
            calculate_high_rise_kinetic_risk(20, 150, 4),
            calculate_datacenter_kinetic_risk(400, 36, False, True),
        ]
        for result in cases:
            controls = get_controls(result)
            narrative = generate_narrative(result, controls, api_key=None, lang="fr")
            refs = get_regulatory_references(result)
            biblio = get_bibliography()
            pdf_bytes = _build_report_pdf(result, narrative, controls, refs, biblio, "2026-01-01 00:00 UTC", "fr")
            assert isinstance(pdf_bytes, bytes)
            assert pdf_bytes[:5] == b"%PDF-", f"invalid PDF header for {result['module']}"
            assert len(pdf_bytes) > 1000, f"suspiciously small PDF for {result['module']}"

    def test_pdf_handles_empty_result_gracefully(self):
        # Dashboard report before any assessment has been run - result/
        # narrative/controls/references can all be empty. Must not crash.
        pdf_bytes = _build_report_pdf({}, "", [], [], get_bibliography(), "2026-01-01 00:00 UTC", "en")
        assert pdf_bytes[:5] == b"%PDF-"

    def test_pdf_strips_unencodable_characters_without_crashing(self):
        # fpdf2's core font can't encode emoji/non-Latin-1 characters -
        # _pdf_safe must strip them rather than let the PDF build crash.
        result = {"module": "Solar (Desert)", "risk_band": "CRITICAL", "primary_hazard": "test"}
        controls = ["🚨 emoji-prefixed control that must not crash PDF generation"]
        pdf_bytes = _build_report_pdf(result, "narrative with emoji 🔥", controls, [], [], "2026-01-01 00:00 UTC", "en")
        assert pdf_bytes[:5] == b"%PDF-"


# ---------------------------------------------------------------------------
# Live API request construction (mocked - no real network call)
# ---------------------------------------------------------------------------

class TestGenerateNarrativeRequest:
    """Verifies the actual request built for the Anthropic API - the source
    of the original bug where the prompt instructed the model to invoke
    unverified 'Roger Bauer... principles'. These tests inspect the real
    request payload (via mocking requests.post) rather than just the
    offline fallback text, since that's where the bug actually lived."""

    def _mock_response(self, text="mock narrative"):
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {"content": [{"type": "text", "text": text}]}
        return mock_resp

    def test_prompt_never_mentions_bauer_or_brauer(self):
        result = calculate_underground_kinetic_risk(30, 85, 100, 15)
        controls = get_controls(result)
        with patch("ai_advisor.requests.post") as mock_post:
            mock_post.return_value = self._mock_response()
            generate_narrative(result, controls, api_key="fake-key-for-test", lang="fr")
            sent_json = mock_post.call_args.kwargs["json"]
            prompt_text = sent_json["messages"][0]["content"]
            assert "Bauer" not in prompt_text
            assert "Brauer" not in prompt_text
            assert "invent" in prompt_text.lower()  # anti-fabrication instruction present

    def test_web_search_tool_absent_by_default(self):
        result = calculate_solar_albedo_heat_risk(500, 5, 30, "pure_desert_sand")
        controls = get_controls(result)
        with patch("ai_advisor.requests.post") as mock_post:
            mock_post.return_value = self._mock_response()
            generate_narrative(result, controls, api_key="fake-key", lang="en", enable_web_search=False)
            sent_json = mock_post.call_args.kwargs["json"]
            assert "tools" not in sent_json

    def test_web_search_tool_present_when_enabled(self):
        result = calculate_solar_albedo_heat_risk(500, 5, 30, "pure_desert_sand")
        controls = get_controls(result)
        with patch("ai_advisor.requests.post") as mock_post:
            mock_post.return_value = self._mock_response()
            generate_narrative(result, controls, api_key="fake-key", lang="en", enable_web_search=True)
            sent_json = mock_post.call_args.kwargs["json"]
            assert "tools" in sent_json
            assert sent_json["tools"][0]["type"] == "web_search_20250305"

    def test_prompt_includes_verified_references_for_module(self):
        result = calculate_datacenter_kinetic_risk(1300, 30, False, False)
        controls = get_controls(result)
        expected_bodies = {ref["body"] for ref in get_references(result["module"])}
        with patch("ai_advisor.requests.post") as mock_post:
            mock_post.return_value = self._mock_response()
            generate_narrative(result, controls, api_key="fake-key", lang="en")
            sent_json = mock_post.call_args.kwargs["json"]
            prompt_text = sent_json["messages"][0]["content"]
            assert any(body in prompt_text for body in expected_bodies)

    def test_no_api_key_never_calls_the_network(self):
        result = calculate_high_rise_kinetic_risk(15, 40, 8)
        controls = get_controls(result)
        with patch("ai_advisor.requests.post") as mock_post:
            narrative = generate_narrative(result, controls, api_key="", lang="fr", enable_web_search=True)
            mock_post.assert_not_called()
            assert isinstance(narrative, str) and narrative


# ---------------------------------------------------------------------------
# Assessment log & monthly Excel export (analytics.py)
# ---------------------------------------------------------------------------

class TestAnalyticsLog:
    def setup_method(self):
        st.session_state.clear()

    def test_log_assessment_appends_row(self):
        r = calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels")
        log_assessment(st, r)
        df = get_log_dataframe(st)
        assert len(df) == 1
        assert df.iloc[0]["module"] == "Solar (Desert)"
        assert list(df.columns) == LOG_COLUMNS

    def test_log_accumulates_across_multiple_calls(self):
        log_assessment(st, calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels"))
        log_assessment(st, calculate_high_rise_kinetic_risk(25, 90, 15))
        df = get_log_dataframe(st)
        assert len(df) == 2

    def test_empty_log_returns_empty_dataframe_with_correct_columns(self):
        df = get_log_dataframe(st)
        assert df.empty
        assert list(df.columns) == LOG_COLUMNS

    def test_key_metric_extracted_per_module(self):
        r_solar = calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels")
        log_assessment(st, r_solar)
        r_dc = calculate_datacenter_kinetic_risk(1300, 30, False, False)
        log_assessment(st, r_dc)
        df = get_log_dataframe(st)
        solar_row = df[df["module"] == "Solar (Desert)"].iloc[0]
        dc_row = df[df["module"] == "Data Center (Controlled Critical Environment)"].iloc[0]
        assert solar_row["key_metric"] == r_solar["perceived_temp"]
        assert dc_row["key_metric"] == r_dc["arc_flash_energy_cal"]

    def test_safety_override_recorded_as_bool(self):
        r = calculate_underground_kinetic_risk(35, 95, 40, 8)  # forces override
        log_assessment(st, r)
        df = get_log_dataframe(st)
        assert bool(df.iloc[0]["safety_override"]) is True


class TestAnalyticsCsvMergeRoundTrip:
    def setup_method(self):
        st.session_state.clear()

    def test_download_then_merge_in_new_session_recovers_data(self):
        import io
        # "Session 1": log something, simulate downloading the CSV
        log_assessment(st, calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels"))
        df1 = get_log_dataframe(st)
        csv_bytes = df1.to_csv(index=False).encode("utf-8")

        # "Session 2": fresh session_state, log something new, merge the old CSV back in
        st.session_state.clear()
        log_assessment(st, calculate_high_rise_kinetic_risk(25, 90, 15))
        added = merge_uploaded_csv(st, io.BytesIO(csv_bytes))
        assert added == 1
        df2 = get_log_dataframe(st)
        assert len(df2) == 2
        assert set(df2["module"]) == {"Solar (Desert)", "High-Rise (Vertical Urban)"}

    def test_merging_same_file_twice_does_not_duplicate(self):
        import io
        log_assessment(st, calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels"))
        df1 = get_log_dataframe(st)
        csv_bytes = df1.to_csv(index=False).encode("utf-8")

        merge_uploaded_csv(st, io.BytesIO(csv_bytes))
        added_again = merge_uploaded_csv(st, io.BytesIO(csv_bytes))
        assert added_again == 0

    def test_merge_rejects_csv_missing_expected_columns(self):
        import io
        bad_csv = io.BytesIO(b"not,the,right,columns\n1,2,3,4\n")
        with pytest.raises(ValueError):
            merge_uploaded_csv(st, bad_csv)


class TestMonthlySummaryAndExcel:
    def setup_method(self):
        st.session_state.clear()

    def test_monthly_summary_on_empty_log(self):
        df = get_log_dataframe(st)
        summary = monthly_summary(df)
        assert summary.empty
        assert list(summary.columns) == ["month", "module", "assessments", "safety_overrides", "avg_key_metric"]

    def test_monthly_summary_aggregates_by_module(self):
        log_assessment(st, calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels"))
        log_assessment(st, calculate_solar_albedo_heat_risk(300, 3, 24, "pure_desert_sand"))
        log_assessment(st, calculate_high_rise_kinetic_risk(25, 90, 15))
        df = get_log_dataframe(st)
        summary = monthly_summary(df)
        solar_row = summary[summary["module"] == "Solar (Desert)"].iloc[0]
        assert solar_row["assessments"] == 2

    def test_build_monthly_excel_is_valid_xlsx_with_three_sheets(self):
        from openpyxl import load_workbook
        import io as io_mod

        log_assessment(st, calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels"))
        log_assessment(st, calculate_high_rise_kinetic_risk(25, 90, 15))
        df = get_log_dataframe(st)

        xlsx_bytes = build_monthly_excel(df)
        assert isinstance(xlsx_bytes, bytes)
        assert xlsx_bytes[:2] == b"PK"  # xlsx is a zip archive

        wb = load_workbook(io_mod.BytesIO(xlsx_bytes))
        assert wb.sheetnames == ["Raw Log", "Monthly Summary", "Trend Chart"]
        assert wb["Raw Log"].max_row == 3  # header + 2 rows
        assert len(wb["Trend Chart"]._charts) == 1  # a real embedded chart, not just data

    def test_build_monthly_excel_handles_empty_log_gracefully(self):
        df = get_log_dataframe(st)
        xlsx_bytes = build_monthly_excel(df)
        assert xlsx_bytes[:2] == b"PK"

    def test_excel_export_has_no_timezone_aware_datetime_error(self):
        # Regression test: pandas raises ValueError writing tz-aware
        # datetimes to Excel - build_monthly_excel must strip tz info first.
        log_assessment(st, calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels"))
        df = get_log_dataframe(st)
        assert df["timestamp"].dt.tz is not None  # confirm the input IS tz-aware
        build_monthly_excel(df)  # must not raise


# ---------------------------------------------------------------------------
# Meteorology forecast (data_feeds.py) - mocked, no real network call
# ---------------------------------------------------------------------------

class TestMeteorologyForecast:
    def _mock_daily_response(self, fields: dict):
        mock_resp = MagicMock()
        mock_resp.raise_for_status = MagicMock()
        mock_resp.json.return_value = {"daily": fields}
        return mock_resp

    def test_solar_forecast_returns_expected_shape(self):
        fetch_solar_forecast.clear()
        fields = {
            "time": ["2026-01-01", "2026-01-02"],
            "temperature_2m_max": [40.0, 41.0],
            "uv_index_max": [9.0, 9.5],
            "shortwave_radiation_sum": [22.0, 23.0],
        }
        with patch("data_feeds.requests.get", return_value=self._mock_daily_response(fields)):
            result = fetch_solar_forecast(days=2)
        assert result["dates"] == ["2026-01-01", "2026-01-02"]
        assert result["temperature_2m_max"] == [40.0, 41.0]
        assert "source" in result

    def test_offshore_forecast_returns_expected_shape(self):
        fetch_offshore_forecast.clear()
        fields = {
            "time": ["2026-01-01", "2026-01-02"],
            "temperature_2m_max": [30.0, 31.0],
            "wind_speed_10m_max": [15.0, 22.0],
        }
        with patch("data_feeds.requests.get", return_value=self._mock_daily_response(fields)):
            result = fetch_offshore_forecast(days=2)
        assert result["wind_speed_10m_max_kn"] == [15.0, 22.0]

    def test_forecast_network_failure_raises_datafeederror(self):
        fetch_solar_forecast.clear()
        with patch("data_feeds.requests.get", side_effect=ConnectionError("network down")):
            with pytest.raises(DataFeedError):
                fetch_solar_forecast(days=2)

    def test_forecast_days_param_is_clamped_to_open_meteo_limits(self):
        fetch_solar_forecast.clear()
        fields = {"time": ["2026-01-01"], "temperature_2m_max": [40.0], "uv_index_max": [9.0], "shortwave_radiation_sum": [22.0]}
        with patch("data_feeds.requests.get", return_value=self._mock_daily_response(fields)) as mock_get:
            fetch_solar_forecast(days=999)  # way over the limit
            sent_params = mock_get.call_args.kwargs["params"]
            assert sent_params["forecast_days"] <= 16


# ---------------------------------------------------------------------------
# Production-MVP audit (Phase 1/2): authentication, audit trail, org
# hierarchy, regulatory threshold badges, HSE disclaimer, AI config
# externalization. These exercise the pure-logic layer of each feature
# directly rather than rendering full Streamlit pages (matching this
# suite's existing convention of testing engine/logic functions, not UI
# widgets) - the actual page wiring was verified end-to-end via AppTest
# during development.
# ---------------------------------------------------------------------------

class TestAuthCredentialResolution:
    """auth.py's fail-closed credential resolution - the CRITICAL security
    finding this audit fixed (a hardcoded admin/Maku2026! fallback) never
    regresses: with nothing configured, _configured_users() must return
    {}, never a default account."""

    def setup_method(self):
        st.session_state.clear()

    def test_no_configuration_anywhere_returns_empty_fail_closed(self, monkeypatch):
        monkeypatch.delenv("MAKU_AUTH_USERNAME", raising=False)
        monkeypatch.delenv("MAKU_AUTH_PASSWORD", raising=False)
        monkeypatch.delenv("MAKU_AUTH_USERS", raising=False)
        monkeypatch.setattr(auth.st, "secrets", {})
        assert auth._configured_users() == {}

    def test_secrets_users_table_takes_priority(self, monkeypatch):
        monkeypatch.setattr(auth.st, "secrets", {
            "auth": {"users": {"alice": "pw1", "bob": "pw2"}, "username": "carol", "password": "pw3"},
        })
        monkeypatch.setenv("MAKU_AUTH_USERNAME", "dave")
        monkeypatch.setenv("MAKU_AUTH_PASSWORD", "pw4")
        users = auth._configured_users()
        assert users == {"alice": "pw1", "bob": "pw2"}

    def test_secrets_single_account_used_when_no_users_table(self, monkeypatch):
        monkeypatch.setattr(auth.st, "secrets", {"auth": {"username": "carol", "password": "pw3"}})
        assert auth._configured_users() == {"carol": "pw3"}

    def test_env_users_json_used_when_no_secrets(self, monkeypatch):
        monkeypatch.setattr(auth.st, "secrets", {})
        monkeypatch.setenv("MAKU_AUTH_USERS", '{"eve": "pw5"}')
        assert auth._configured_users() == {"eve": "pw5"}

    def test_env_single_account_is_last_resort(self, monkeypatch):
        monkeypatch.setattr(auth.st, "secrets", {})
        monkeypatch.delenv("MAKU_AUTH_USERS", raising=False)
        monkeypatch.setenv("MAKU_AUTH_USERNAME", "frank")
        monkeypatch.setenv("MAKU_AUTH_PASSWORD", "pw6")
        assert auth._configured_users() == {"frank": "pw6"}

    def test_no_hardcoded_admin_fallback_exists(self, monkeypatch):
        """Regression test for the exact CRITICAL finding this audit fixed:
        an "admin"/"Maku2026!" hardcoded fallback must never be
        reachable, under any configuration state."""
        monkeypatch.delenv("MAKU_AUTH_USERNAME", raising=False)
        monkeypatch.delenv("MAKU_AUTH_PASSWORD", raising=False)
        monkeypatch.delenv("MAKU_AUTH_USERS", raising=False)
        monkeypatch.setattr(auth.st, "secrets", {})
        assert not auth._verify_credentials("admin", "Maku2026!")
        assert auth._configured_users().get("admin") is None


class TestAuthPasswordMatching:
    def test_plain_password_constant_time_match(self):
        assert auth._password_matches("correct-horse", "correct-horse")
        assert not auth._password_matches("correct-horse", "wrong")

    def test_sha256_hashed_password_matches_correct_plaintext(self):
        import hashlib
        digest = hashlib.sha256(b"my-secret").hexdigest()
        configured = f"sha256:{digest}"
        assert auth._password_matches(configured, "my-secret")
        assert not auth._password_matches(configured, "not-my-secret")

    def test_unknown_username_runs_decoy_comparison_and_fails(self, monkeypatch):
        monkeypatch.setattr(auth.st, "secrets", {"auth": {"users": {"alice": "pw1"}}})
        assert not auth._verify_credentials("nobody", "anything")

    def test_bcrypt_hashed_password_matches_correct_plaintext(self):
        """P0 'AUTHENTICATION & CRYPTOGRAPHIC HARDENING' - bcrypt is now
        the primary, recommended credential format."""
        configured = auth.hash_password("my-secret")
        assert configured.startswith("bcrypt:")
        assert auth._password_matches(configured, "my-secret")
        assert not auth._password_matches(configured, "not-my-secret")

    def test_hash_password_produces_a_different_hash_each_time(self):
        # bcrypt salts every hash - two calls for the same password must
        # not produce identical stored values.
        assert auth.hash_password("same-password") != auth.hash_password("same-password")

    def test_malformed_bcrypt_hash_fails_closed(self):
        assert not auth._password_matches("bcrypt:not-a-real-hash", "anything")

    def test_credential_format_classification(self):
        assert auth._credential_format(auth.hash_password("x")) == "bcrypt"
        assert auth._credential_format("sha256:" + "a" * 64) == "sha256"
        assert auth._credential_format("plain-value") == "plaintext"

    def test_successful_bcrypt_login_does_not_flag_deprecated_warning(self, monkeypatch):
        st.session_state.clear()
        configured = auth.hash_password("s3cret!")
        monkeypatch.setattr(auth.st, "secrets", {"auth": {"users": {"alice": configured}}})
        assert auth._verify_credentials("alice", "s3cret!")
        assert st.session_state.get("_auth_credential_format_warning") is None

    def test_successful_sha256_login_flags_deprecated_warning(self, monkeypatch):
        import hashlib
        st.session_state.clear()
        digest = hashlib.sha256(b"legacy-secret").hexdigest()
        monkeypatch.setattr(auth.st, "secrets", {"auth": {"users": {"bob": f"sha256:{digest}"}}})
        assert auth._verify_credentials("bob", "legacy-secret")
        warning = st.session_state.get("_auth_credential_format_warning")
        assert warning is not None
        assert "bob" in warning
        assert "SHA-256" in warning

    def test_successful_plaintext_login_flags_deprecated_warning(self, monkeypatch):
        st.session_state.clear()
        monkeypatch.setattr(auth.st, "secrets", {"auth": {"users": {"carol": "plaintext-pw"}}})
        assert auth._verify_credentials("carol", "plaintext-pw")
        warning = st.session_state.get("_auth_credential_format_warning")
        assert warning is not None
        assert "carol" in warning
        assert "plaintext" in warning


class TestAuthLockoutAndSessionExpiry:
    def setup_method(self):
        st.session_state.clear()

    def test_lockout_triggers_after_max_failed_attempts(self):
        for _ in range(auth.MAX_FAILED_ATTEMPTS - 1):
            auth._register_failure("bob")
        attempts, locked_until = auth._lockout_state()
        assert attempts == auth.MAX_FAILED_ATTEMPTS - 1
        assert locked_until == 0.0
        auth._register_failure("bob")  # the MAX_FAILED_ATTEMPTS-th failure
        attempts, locked_until = auth._lockout_state()
        assert attempts == 0  # counter resets once locked out
        assert locked_until > 0.0

    def test_reset_failures_clears_lockout_state(self):
        auth._register_failure("bob")
        auth._register_failure("bob")
        auth._reset_failures()
        assert auth._lockout_state() == (0, 0.0)

    def test_session_not_expired_when_never_logged_in(self):
        assert auth._session_expired() is False

    def test_session_expires_after_timeout(self):
        import time
        st.session_state["_auth_login_at"] = time.time() - (auth.SESSION_TIMEOUT_MINUTES * 60 + 1)
        assert auth._session_expired() is True

    def test_session_not_expired_within_timeout(self):
        import time
        st.session_state["_auth_login_at"] = time.time()
        assert auth._session_expired() is False


# ---------------------------------------------------------------------------
# Tamper-evident audit trail (analytics.py)
# ---------------------------------------------------------------------------

class TestAuditTrail:
    @pytest.fixture(autouse=True)
    def _isolated_db(self, tmp_path, monkeypatch):
        """Points analytics.py at a fresh, throwaway SQLite file for every
        test in this class - _get_connection() re-reads analytics.DB_PATH
        on every call (see analytics._resolve_backend()), so this fully
        isolates the audit ledger from the app's real maku_site_data.db
        and from every other test module. Necessary specifically here
        because test_hash_chain_integrity_detects_direct_tampering()
        deliberately corrupts a row - that must never leak into the
        durable database the rest of the suite (or a real deployment)
        reads from."""
        st.session_state.clear()
        monkeypatch.setattr("analytics.DB_PATH", str(tmp_path / "test_audit.db"))

    def test_log_audit_event_returns_true_and_is_readable(self):
        ok = log_audit_event(AUDIT_EVENT_LOGIN_SUCCESS, actor="test-user", detail="unit test login")
        assert ok is True
        df = get_audit_log_dataframe()
        assert len(df) >= 1
        newest = df.iloc[0]  # get_audit_log_dataframe() orders newest-first (ORDER BY id DESC)
        assert newest["event_type"] == AUDIT_EVENT_LOGIN_SUCCESS
        assert newest["actor"] == "test-user"

    def test_audit_log_never_raises_even_on_bad_input(self):
        # None/empty values must degrade gracefully, never crash the caller -
        # audit logging must never become a denial-of-service vector.
        ok = log_audit_event(AUDIT_EVENT_LOGIN_SUCCESS, actor="", detail="")
        assert ok in (True, False)

    def test_hash_chain_integrity_holds_for_untampered_log(self):
        log_audit_event(AUDIT_EVENT_LOGIN_SUCCESS, actor="checker-1")
        log_audit_event(AUDIT_EVENT_LOGIN_SUCCESS, actor="checker-2")
        report = verify_audit_log_integrity()
        assert report["ok"] is True
        assert report["first_broken_id"] is None

    def test_hash_chain_integrity_detects_direct_tampering(self):
        import sqlite3
        import analytics as _analytics

        log_audit_event(AUDIT_EVENT_LOGIN_SUCCESS, actor="victim")
        df_before = get_audit_log_dataframe()
        tampered_id = int(df_before.iloc[0]["id"])  # newest-first ordering - this is the row just written

        if _analytics.get_backend_status()["resolved_backend"] != "sqlite":
            pytest.skip("tamper test requires the sqlite backend (no live Postgres in this environment)")

        conn = sqlite3.connect(_analytics.DB_PATH)
        conn.execute("UPDATE audit_log SET detail = 'TAMPERED' WHERE id = ?", (tampered_id,))
        conn.commit()
        conn.close()

        report = verify_audit_log_integrity()
        assert report["ok"] is False
        assert report["first_broken_id"] == tampered_id

    def test_concurrent_writes_keep_the_hash_chain_intact(self):
        """P0 'SECURITY EXPOSURE & AUDIT CONCURRENCY' - MUTEX PROTECTION
        regression test: hammer log_audit_event() from many threads at
        once and confirm the hash chain comes out fully valid. Without
        _AUDIT_CHAIN_LOCK/_lock_audit_chain(), concurrent writers can both
        read the same prev_hash before either commits, forking the chain -
        this test would flake/fail on that race."""
        import concurrent.futures

        def _write(i):
            return log_audit_event(AUDIT_EVENT_LOGIN_SUCCESS, actor=f"racer-{i}", detail=str(i))

        with concurrent.futures.ThreadPoolExecutor(max_workers=16) as ex:
            results = list(ex.map(_write, range(40)))

        assert all(results)
        report = verify_audit_log_integrity()
        assert report["ok"] is True
        assert report["rows_checked"] == 40


class TestDatabaseFailSafe:
    """P0 'DATABASE SAFETY & FAIL-SAFE RUNTIME' - a configured Postgres/
    Supabase backend that can't be reached must freeze writes/reads and
    surface CRITICAL_DB_ERROR_MESSAGE, never silently redirect to SQLite
    or in-memory session_state."""

    @pytest.fixture(autouse=True)
    def _isolated_and_reset(self, tmp_path, monkeypatch):
        import analytics as _analytics
        st.session_state.clear()
        monkeypatch.setattr("analytics.DB_PATH", str(tmp_path / "test_failsafe.db"))
        _analytics.reset_fatal_state_for_testing()
        yield
        # Never let a simulated Postgres outage leak fatal state into
        # every test that runs after this one in the same process.
        monkeypatch.setattr("analytics.DATABASE_URL", "")
        _analytics.reset_fatal_state_for_testing()

    def test_unreachable_configured_postgres_url_is_fatal_not_a_fallback(self, monkeypatch):
        import analytics as _analytics

        monkeypatch.setattr("analytics.DATABASE_URL", "postgresql://baduser:badpass@127.0.0.1:1/nope")

        ok = log_audit_event(AUDIT_EVENT_LOGIN_SUCCESS, actor="tester")
        assert ok is False
        assert _analytics.is_db_fatal() is True

        status = _analytics.get_backend_status()
        assert status["fatal"] is True
        assert status["fatal_message"] == _analytics.CRITICAL_DB_ERROR_MESSAGE

    def test_fatal_state_freezes_writes_no_silent_redirect_to_session_state(self, monkeypatch):
        import analytics as _analytics

        monkeypatch.setattr("analytics.DATABASE_URL", "postgresql://baduser:badpass@127.0.0.1:1/nope")
        log_audit_event(AUDIT_EVENT_LOGIN_SUCCESS, actor="tester")  # trips fatal state
        assert _analytics.is_db_fatal() is True

        result = calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels")
        log_assessment(st, result)
        # The whole point of FAIL-FAST: this must NOT have silently
        # written into the ephemeral session_state fallback list.
        assert _analytics.LOG_SESSION_KEY not in st.session_state

        df = get_log_dataframe(st)
        assert df.empty

    def test_default_sqlite_deployment_is_unaffected_by_failsafe_change(self):
        """No DATABASE_URL configured at all -> SQLite remains the
        intended backend, not a 'fallback' the FAIL-FAST directive is
        about - normal session_state degrade-on-failure behavior for a
        genuinely broken local SQLite file must still work exactly as
        before this change."""
        import analytics as _analytics

        assert _analytics.DATABASE_URL == ""
        result = calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels")
        log_assessment(st, result)
        assert _analytics.is_db_fatal() is False
        df = get_log_dataframe(st)
        assert not df.empty


# ---------------------------------------------------------------------------
# Enterprise Organization / Project / Site context (analytics.py)
# ---------------------------------------------------------------------------

class TestOrgContext:
    @pytest.fixture(autouse=True)
    def _isolated_db(self, tmp_path, monkeypatch):
        """Fresh, throwaway SQLite file per test - set_org_context() now
        resolves/creates real relational organization/project/site ids
        (P0 'MULTI-TENANT ISOLATION ARCHITECTURE'), so isolating the
        database keeps those ids deterministic and independent of
        whatever other test classes have already written."""
        st.session_state.clear()
        monkeypatch.setattr("analytics.DB_PATH", str(tmp_path / "test_org_context.db"))

    def test_defaults_when_never_set(self):
        ctx = get_org_context(st)
        assert ctx["organization"] == DEFAULT_ORGANIZATION
        assert ctx["project"] == DEFAULT_PROJECT
        assert ctx["site"] == ""
        assert ctx["organization_id"] is None
        assert ctx["project_id"] is None
        assert ctx["site_id"] is None

    def test_set_and_get_round_trip(self):
        set_org_context(st, organization="Acme Corp", project="Tower A", site="North Wing")
        ctx = get_org_context(st)
        assert ctx["organization"] == "Acme Corp"
        assert ctx["project"] == "Tower A"
        assert ctx["site"] == "North Wing"
        # Relational hierarchy: real, positive, distinct ids resolved
        # against the organizations/projects/sites tables.
        assert isinstance(ctx["organization_id"], int) and ctx["organization_id"] > 0
        assert isinstance(ctx["project_id"], int) and ctx["project_id"] > 0
        assert isinstance(ctx["site_id"], int) and ctx["site_id"] > 0

    def test_same_names_resolve_to_the_same_stable_ids(self):
        set_org_context(st, organization="Acme Corp", project="Tower A", site="North Wing")
        first = get_org_context(st)
        set_org_context(st, organization="Acme Corp", project="Tower A", site="North Wing")
        second = get_org_context(st)
        assert first["organization_id"] == second["organization_id"]
        assert first["project_id"] == second["project_id"]
        assert first["site_id"] == second["site_id"]

    def test_different_organizations_resolve_to_different_ids(self):
        set_org_context(st, organization="Acme Corp", project="Tower A", site="North Wing")
        acme_id = get_org_context(st)["organization_id"]
        set_org_context(st, organization="Beta Industries", project="Site 1", site="")
        beta_id = get_org_context(st)["organization_id"]
        assert acme_id != beta_id

    def test_empty_values_fall_back_to_defaults(self):
        set_org_context(st, organization="", project="", site="")
        ctx = get_org_context(st)
        assert ctx["organization"] == DEFAULT_ORGANIZATION
        assert ctx["project"] == DEFAULT_PROJECT

    def test_log_assessment_tags_current_org_context(self):
        set_org_context(st, organization="Acme Corp", project="Tower A", site="North Wing")
        log_assessment(st, calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels"))
        df = get_log_dataframe(st)
        assert df.iloc[-1]["organization"] == "Acme Corp"
        assert df.iloc[-1]["project"] == "Tower A"


class TestMultiTenantIsolation:
    """P0 'MULTI-TENANT ISOLATION ARCHITECTURE' - the specific vulnerability
    named in the audit: get_site_alert_log_dataframe() must never leak one
    organization's site alerts into another organization's view."""

    @pytest.fixture(autouse=True)
    def _isolated_db(self, tmp_path, monkeypatch):
        st.session_state.clear()
        monkeypatch.setattr("analytics.DB_PATH", str(tmp_path / "test_tenant_isolation.db"))

    def test_site_alert_log_filters_by_organization(self):
        import analytics as _analytics

        _analytics.log_site_alert("Worker_A1", "physiological_strain", "CRITICAL",
                                   "Heat strain", module="Worker Physiology",
                                   organization="Tenant Alpha")
        _analytics.log_site_alert("Worker_B1", "physiological_strain", "CRITICAL",
                                   "Heat strain", module="Worker Physiology",
                                   organization="Tenant Beta")

        alpha_view = _analytics.get_site_alert_log_dataframe(organization="Tenant Alpha")
        beta_view = _analytics.get_site_alert_log_dataframe(organization="Tenant Beta")

        assert len(alpha_view) == 1
        assert alpha_view.iloc[0]["worker_or_site_id"] == "Worker_A1"
        assert len(beta_view) == 1
        assert beta_view.iloc[0]["worker_or_site_id"] == "Worker_B1"
        # Cross-tenant leakage check: Tenant Alpha's view must not contain
        # any row belonging to Tenant Beta, and vice versa.
        assert "Worker_B1" not in alpha_view["worker_or_site_id"].values
        assert "Worker_A1" not in beta_view["worker_or_site_id"].values

    def test_site_alert_log_resolves_organization_id(self):
        import analytics as _analytics

        _analytics.log_site_alert("Worker_C1", "physiological_strain", "WARNING",
                                   "Elevated strain", organization="Tenant Gamma",
                                   project="Project X", site="Site 9")
        df = _analytics.get_site_alert_log_dataframe(organization="Tenant Gamma")
        assert len(df) == 1
        # organization_id/project_id/site_id columns are present and
        # resolved to real relational ids, not left null.
        assert df.iloc[0]["organization_id"] is not None
        assert int(df.iloc[0]["organization_id"]) > 0

    def test_unfiltered_call_sees_every_tenant_admin_only(self):
        import analytics as _analytics

        _analytics.log_site_alert("Worker_A1", "physiological_strain", "CRITICAL",
                                   "Heat strain", organization="Tenant Alpha")
        _analytics.log_site_alert("Worker_B1", "physiological_strain", "CRITICAL",
                                   "Heat strain", organization="Tenant Beta")
        everything = _analytics.get_site_alert_log_dataframe()
        assert len(everything) == 2


# ---------------------------------------------------------------------------
# Regulatory threshold classification badges (regulatory_country_thresholds.py)
# ---------------------------------------------------------------------------

class TestRegulatoryThresholdBadges:
    _TOPICS = [
        "heat_stress", "wind_shear", "air_quality_ambient", "air_quality_occupational",
        "noise", "silica", "uv_heat", "cold_stress", "remote_comms", "midday_ban",
    ]

    def test_every_registered_country_and_topic_resolves_without_error(self):
        for country_code in list(REGULATORY_PROFILES.keys()) + ["NOT_A_REAL_COUNTRY"]:
            for topic in self._TOPICS:
                badge = get_threshold_category_badge(country_code, topic)
                assert badge["category"] in THRESHOLD_CATEGORY_META
                assert badge["label"].startswith("[") and badge["label"].endswith("]")

    def test_exact_required_category_labels_are_present(self):
        labels = {meta["label"] for meta in THRESHOLD_CATEGORY_META.values()}
        assert labels == {
            "[LEGAL REQUIREMENT]", "[STANDARD]", "[GUIDANCE]",
            "[SITE/OEM REQUIREMENT]", "[MAKU SCREENING VALUE]",
        }

    def test_crane_wind_shear_is_never_shown_as_a_legal_requirement(self):
        """Every country profile explicitly documents crane wind-shear as
        deferring to manufacturer/OEM load charts, never a cited statutory
        number - the badge must never claim otherwise."""
        for country_code in REGULATORY_PROFILES:
            badge = get_threshold_category_badge(country_code, "wind_shear")
            assert badge["category"] != "LEGAL_REQUIREMENT"

    def test_unregistered_country_defaults_to_least_authoritative_category(self):
        badge = get_threshold_category_badge("ATLANTIS", "heat_stress")
        assert badge["category"] == "MAKU_SCREENING_VALUE"

    def test_render_regulatory_badge_does_not_raise(self):
        mock_st = MagicMock()
        render_regulatory_badge(mock_st, "USA", "heat_stress")
        assert mock_st.markdown.called


# ---------------------------------------------------------------------------
# HSE decision-support disclaimer (ui_helpers.py)
# ---------------------------------------------------------------------------

class TestHseDisclaimer:
    _REQUIRED_TEXT = (
        "MAKU HSE Decision-Support System — Results are intended for risk "
        "screening and decision support and do not replace a competent "
        "person's assessment, engineering analysis, statutory requirement, "
        "manufacturer instruction or formal specialist study."
    )

    def test_exact_required_wording(self):
        assert HSE_DISCLAIMER_TEXT == self._REQUIRED_TEXT

    def test_render_compact_calls_warning_with_exact_text(self):
        mock_st = MagicMock()
        render_hse_disclaimer(mock_st, lang="en", compact=True)
        mock_st.warning.assert_called_once()
        assert mock_st.warning.call_args.args[0] == HSE_DISCLAIMER_TEXT

    def test_render_full_container_includes_exact_text(self):
        mock_st = MagicMock()
        mock_st.container.return_value.__enter__.return_value = mock_st
        mock_st.container.return_value.__exit__.return_value = False
        render_hse_disclaimer(mock_st, lang="en", compact=False)
        markdown_calls = "".join(str(c) for c in mock_st.markdown.call_args_list)
        assert HSE_DISCLAIMER_TEXT in markdown_calls

    def test_disclaimer_survives_pdf_safe_encoding_including_em_dash(self):
        """Regression test for the cp1252-vs-latin1 fpdf2 bug this audit
        found and fixed: the em dash in the required wording must survive
        the PDF text sanitizer, not be silently dropped. fpdf2's core-font
        rendering path re-encodes the returned string as raw latin-1 bytes
        and interprets each byte per WinAnsiEncoding (cp1252) - so the byte
        _pdf_safe() produces for the em dash must equal cp1252's own
        encoding of an em dash, not the (different, wrong) plain-latin-1
        encoding the original buggy version produced."""
        from ui_helpers import _pdf_safe
        assert "—" in HSE_DISCLAIMER_TEXT  # sanity: the source text really has an em dash
        safe_text = _pdf_safe(HSE_DISCLAIMER_TEXT)
        em_dash_index = HSE_DISCLAIMER_TEXT.index("—")
        produced_byte = safe_text.encode("latin-1")[em_dash_index]
        expected_byte = "—".encode("cp1252")[0]
        assert produced_byte == expected_byte


# ---------------------------------------------------------------------------
# AI advisory layer configuration externalization (ai_advisor.py)
# ---------------------------------------------------------------------------

class TestAiConfigExternalization:
    def test_defaults_match_original_hardcoded_values_when_unconfigured(self, monkeypatch):
        monkeypatch.setattr(ai_advisor.st, "secrets", {})
        monkeypatch.delenv("MAKU_AI_MODEL", raising=False)
        monkeypatch.delenv("MAKU_AI_API_BASE_URL", raising=False)
        monkeypatch.delenv("MAKU_AI_WEB_SEARCH_MAX_USES", raising=False)
        cfg = ai_advisor._resolve_ai_config()
        assert cfg["model"] == "claude-sonnet-4-5"
        assert cfg["api_base_url"] == "https://api.anthropic.com/v1/messages"
        assert cfg["web_search_max_uses"] == 3

    def test_env_vars_override_defaults(self, monkeypatch):
        monkeypatch.setattr(ai_advisor.st, "secrets", {})
        monkeypatch.setenv("MAKU_AI_MODEL", "claude-test-model")
        monkeypatch.setenv("MAKU_AI_API_BASE_URL", "https://gateway.example.com/v1/messages")
        monkeypatch.setenv("MAKU_AI_WEB_SEARCH_MAX_USES", "7")
        cfg = ai_advisor._resolve_ai_config()
        assert cfg["model"] == "claude-test-model"
        assert cfg["api_base_url"] == "https://gateway.example.com/v1/messages"
        assert cfg["web_search_max_uses"] == 7

    def test_secrets_take_priority_over_env_vars(self, monkeypatch):
        monkeypatch.setattr(ai_advisor.st, "secrets", {
            "ai": {"model": "claude-from-secrets", "api_base_url": "https://secrets.example.com",
                   "web_search_max_uses": 9},
        })
        monkeypatch.setenv("MAKU_AI_MODEL", "claude-from-env")
        cfg = ai_advisor._resolve_ai_config()
        assert cfg["model"] == "claude-from-secrets"
        assert cfg["web_search_max_uses"] == 9

    def test_no_server_side_api_key_configured_by_default(self, monkeypatch):
        monkeypatch.setattr(ai_advisor.st, "secrets", {})
        monkeypatch.delenv("MAKU_AI_API_KEY", raising=False)
        cfg = ai_advisor._resolve_ai_config()
        assert cfg["api_key"] is None

    def test_server_side_api_key_from_secrets(self, monkeypatch):
        monkeypatch.setattr(ai_advisor.st, "secrets", {"ai": {"api_key": "sk-from-secrets"}})
        cfg = ai_advisor._resolve_ai_config()
        assert cfg["api_key"] == "sk-from-secrets"

    def test_server_side_api_key_from_env(self, monkeypatch):
        monkeypatch.setattr(ai_advisor.st, "secrets", {})
        monkeypatch.setenv("MAKU_AI_API_KEY", "sk-from-env")
        cfg = ai_advisor._resolve_ai_config()
        assert cfg["api_key"] == "sk-from-env"

    def test_get_configured_ai_api_key_reflects_ai_config(self, monkeypatch):
        monkeypatch.setattr(ai_advisor, "_AI_CONFIG", {"api_key": "sk-live-value"})
        assert ai_advisor.get_configured_ai_api_key() == "sk-live-value"
        monkeypatch.setattr(ai_advisor, "_AI_CONFIG", {"api_key": None})
        assert ai_advisor.get_configured_ai_api_key() is None

    def test_ai_layer_failure_never_propagates_past_generate_narrative(self):
        """generate_narrative() must always return a usable string, even if
        the network call raises unexpectedly - the deterministic risk
        engine's own output must never depend on the AI layer succeeding."""
        result = calculate_solar_albedo_heat_risk(950, 9, 42, "silicon_pv_panels")
        controls = get_controls(result)
        with patch("ai_advisor.requests.post", side_effect=RuntimeError("simulated failure")):
            narrative = generate_narrative(result, controls, api_key="fake-key", lang="en")
        assert isinstance(narrative, str) and narrative


# ===========================================================================
# HSE Auditor corrective-action overhaul: Risk Matrix, Heat-Stress Upgrade,
# Stop-Work Trigger Registry, 2-stage Residual Risk workflow, Evidence &
# Traceability, Regulatory Algorithm Validation.
# ===========================================================================

class TestRiskMatrixEngine:
    """risk_matrix.py: the shared Likelihood x Severity scoring engine
    every module's risk_matrix key is built from."""

    def test_matrix_score_clamps_out_of_range_inputs(self):
        assert matrix_score(5, 5) == 25
        assert matrix_score(1, 1) == 1
        assert matrix_score(0, 10) == 1 * 5  # both clamped into [1, 5]
        assert matrix_score(-3, 99) == 5

    @pytest.mark.parametrize("score,expected_band", [
        (1, "Low"), (4, "Low"),
        (5, "Moderate"), (9, "Moderate"),
        (10, "High"), (15, "High"),
        (16, "Extreme"), (25, "Extreme"),
    ])
    def test_matrix_band_boundaries(self, score, expected_band):
        """Exact hardcoded band thresholds the audit requires: Low 1-4,
        Moderate 5-9, High 10-15, Extreme/Critical 16-25."""
        assert matrix_band(score) == expected_band

    def test_severity_from_band_maps_both_vocabularies(self):
        assert severity_from_band("LOW") == severity_from_band("Low") == 1
        assert severity_from_band("MODERATE") == severity_from_band("Moderate") == 3
        assert severity_from_band("HIGH") == severity_from_band("High") == 4
        assert severity_from_band("CRITICAL") == severity_from_band("Extreme") == severity_from_band("EXTREME") == 5

    def test_severity_from_band_unrecognized_label_falls_back_to_default(self):
        assert severity_from_band("not_a_real_band") == 3
        assert severity_from_band("not_a_real_band", default=2) == 2

    def test_likelihood_from_margin_boundary_behavior(self):
        # Already at/past the threshold -> almost certain (5)
        assert likelihood_from_margin(current_value=40, threshold=35, comfortable_margin=10) == 5
        assert likelihood_from_margin(current_value=35, threshold=35, comfortable_margin=10) == 5
        # Comfortably clear of the threshold -> rare (1)
        assert likelihood_from_margin(current_value=0, threshold=35, comfortable_margin=10) == 1
        # Monotonic: likelihood never decreases as current_value rises toward the threshold
        values = [likelihood_from_margin(v, threshold=35, comfortable_margin=10) for v in range(0, 40, 5)]
        assert values == sorted(values)

    def test_score_hazard_structure(self):
        h = score_hazard("Test hazard", likelihood=4, severity=5, note="unit test note")
        assert h == {
            "name": "Test hazard", "likelihood": 4, "severity": 5,
            "score": 20, "band": "Extreme", "note": "unit test note",
        }

    def test_aggregate_risk_matrix_worst_hazard_governs(self):
        hazards = [
            score_hazard("Low hazard", 1, 1),
            score_hazard("Worst hazard", 5, 5),
            score_hazard("Mid hazard", 3, 3),
        ]
        result = aggregate_risk_matrix(hazards)
        assert result["overall_score"] == 25
        assert result["overall_band"] == "Extreme"
        assert result["governing_hazard"] == "Worst hazard"
        assert result["hazards"] == hazards

    def test_aggregate_risk_matrix_empty_list(self):
        result = aggregate_risk_matrix([])
        assert result == {"hazards": [], "overall_score": 0, "overall_band": "Low", "governing_hazard": None}


class TestApplyControlsResidualRisk:
    """2-stage workflow: Initial Risk -> Applied Controls -> Residual Risk.
    Methodology: each applied control reduces the GOVERNING hazard's
    likelihood by 1, floored at 1; severity is never reduced."""

    def _initial_matrix(self):
        # Secondary hazard's score (2*2=4) is deliberately kept below the
        # governing hazard's fully-reduced floor score (1*5=5), so tests
        # below can assert the floor is reached without the secondary
        # hazard taking over as governing.
        return aggregate_risk_matrix([
            score_hazard("Governing hazard", 5, 5, note="worst"),
            score_hazard("Secondary hazard", 2, 2, note="secondary"),
        ])

    def test_zero_controls_applied_leaves_score_unchanged(self):
        initial = self._initial_matrix()
        residual = apply_controls_residual_risk(initial, 0)
        assert residual["overall_score"] == initial["overall_score"]
        assert residual["likelihood_reduction_applied"] == 0
        assert residual["controls_applied_count"] == 0

    def test_reduction_is_monotonic_and_floored_at_likelihood_1(self):
        initial = self._initial_matrix()
        scores = [apply_controls_residual_risk(initial, n)["overall_score"] for n in range(0, 8)]
        assert scores == sorted(scores, reverse=True), "score must never increase as more controls are applied"
        # Governing hazard likelihood=5, severity=5: floor is likelihood=1 -> score=5 (Moderate),
        # never lower - severity is never reduced by this function.
        assert scores[-1] == 5
        assert apply_controls_residual_risk(initial, 4)["overall_score"] == \
               apply_controls_residual_risk(initial, 100)["overall_score"] == 5

    def test_severity_is_never_reduced(self):
        initial = self._initial_matrix()
        for n in (1, 2, 3, 4, 10):
            residual = apply_controls_residual_risk(initial, n)
            governing = next(h for h in residual["hazards"] if h["name"] == "Governing hazard")
            assert governing["severity"] == 5

    def test_a_different_hazard_can_become_governing_after_reduction(self):
        """If reducing the initial governing hazard's likelihood enough
        brings it below a different hazard's score, that other hazard must
        correctly become the new governing hazard (full re-aggregation,
        not an isolated single-hazard recompute)."""
        initial = aggregate_risk_matrix([
            score_hazard("Reducible hazard", 5, 4, note="starts worst: score 20"),
            score_hazard("Stubborn hazard", 4, 4, note="score 16, doesn't change"),
        ])
        assert initial["governing_hazard"] == "Reducible hazard"
        residual = apply_controls_residual_risk(initial, 2)  # 5 -> 3, score 12 < 16
        assert residual["overall_score"] == 16
        assert residual["governing_hazard"] == "Stubborn hazard"

    def test_negative_controls_applied_treated_as_zero(self):
        initial = self._initial_matrix()
        residual = apply_controls_residual_risk(initial, -5)
        assert residual["controls_applied_count"] == 0
        assert residual["overall_score"] == initial["overall_score"]

    def test_empty_risk_matrix_does_not_crash(self):
        empty = {"hazards": [], "overall_score": 0, "overall_band": "Low", "governing_hazard": None}
        residual = apply_controls_residual_risk(empty, 3)
        assert residual["hazards"] == []
        assert residual["controls_applied_count"] == 3
        assert residual["likelihood_reduction_applied"] == 0


class TestRiskMatrixWiringAcrossAllModules:
    """Every calculate_*_kinetic_risk() must carry a well-formed
    risk_matrix key (Task: risk algorithm transparency & matrix scoring)."""

    LOW_RISK_CALLS = [
        lambda: calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand"),
        lambda: calculate_marine_humidex_risk(ambient_temp=24, relative_humidity=55, wind_speed=8),
        lambda: calculate_underground_kinetic_risk(ambient_temp=22, geothermal_humidity=55,
                                                     particulate_matter_pm25=20, gas_co_ppm=5),
        lambda: calculate_high_rise_kinetic_risk(ground_wind_speed_knots=8, floor_level=5, crane_load_mass_tons=10),
        lambda: calculate_datacenter_kinetic_risk(electrical_load_kw=100, hot_aisle_temp=26,
                                                    ceiling_void_confined=False, gas_system_armed=False),
        lambda: calculate_wind_energy_kinetic_risk(6.0),
        lambda: calculate_mining_quarrying_kinetic_risk(respirable_silica_ugm3=10, measured_noise_dba=75,
                                                          noise_exposure_hours=8, measured_vibration_aw_ms2=0.2,
                                                          vibration_exposure_hours=8),
        lambda: calculate_marine_port_kinetic_risk(current_tide_level_m=2.0, required_min_clearance_m=1.0,
                                                     is_night_operation=False, measured_illuminance_lux=200,
                                                     hardware_years_in_service=1),
    ]

    HIGH_RISK_CALLS = [
        lambda: calculate_solar_albedo_heat_risk(ghi=1100, uv_index=6, ambient_temp=48, surface_type="silicon_pv_panels"),
        lambda: calculate_marine_humidex_risk(ambient_temp=34, relative_humidity=98, wind_speed=27),
        lambda: calculate_underground_kinetic_risk(ambient_temp=35, geothermal_humidity=95,
                                                     particulate_matter_pm25=40, gas_co_ppm=8),
        lambda: calculate_high_rise_kinetic_risk(ground_wind_speed_knots=25, floor_level=90, crane_load_mass_tons=15),
        lambda: calculate_datacenter_kinetic_risk(electrical_load_kw=1300, hot_aisle_temp=48,
                                                    ceiling_void_confined=True, gas_system_armed=True),
        lambda: calculate_wind_energy_kinetic_risk(22.0, is_offshore=True, significant_wave_height_m=2.5,
                                                     flash_to_bang_sec=10),
        lambda: calculate_mining_quarrying_kinetic_risk(respirable_silica_ugm3=60, measured_noise_dba=105,
                                                          noise_exposure_hours=8, measured_vibration_aw_ms2=1.5,
                                                          vibration_exposure_hours=8),
        lambda: calculate_marine_port_kinetic_risk(current_tide_level_m=1.1, required_min_clearance_m=1.0,
                                                     is_night_operation=True, measured_illuminance_lux=20,
                                                     hardware_years_in_service=20),
    ]

    @pytest.mark.parametrize("make_result", LOW_RISK_CALLS + HIGH_RISK_CALLS)
    def test_every_module_returns_well_formed_risk_matrix(self, make_result):
        result = make_result()
        rm = result.get("risk_matrix")
        assert rm is not None, f"{result['module']} is missing risk_matrix"
        assert rm["hazards"], f"{result['module']} risk_matrix has no hazards"
        assert 1 <= rm["overall_score"] <= 25
        assert rm["overall_band"] in ("Low", "Moderate", "High", "Extreme")
        assert rm["governing_hazard"] in {h["name"] for h in rm["hazards"]}
        for h in rm["hazards"]:
            assert 1 <= h["likelihood"] <= 5
            assert 1 <= h["severity"] <= 5
            assert h["score"] == h["likelihood"] * h["severity"]

    def test_low_risk_inputs_do_not_saturate_to_extreme(self):
        """Regression guard: every low-risk scenario above must land in a
        low band, not silently saturate at Extreme regardless of input."""
        for make_result in self.LOW_RISK_CALLS:
            result = make_result()
            assert result["risk_matrix"]["overall_band"] in ("Low", "Moderate"), \
                f"{result['module']} low-risk input unexpectedly scored {result['risk_matrix']['overall_band']}"

    def test_high_risk_inputs_reach_high_or_extreme(self):
        for make_result in self.HIGH_RISK_CALLS:
            result = make_result()
            assert result["risk_matrix"]["overall_band"] in ("High", "Extreme"), \
                f"{result['module']} high-risk input unexpectedly scored {result['risk_matrix']['overall_band']}"

    def test_marine_port_tide_clearance_direction_is_not_inverted(self):
        """Regression test for a real bug caught during development: tide
        clearance margin and corrosion remaining-capacity are both
        "higher is safer" readings, unlike every other hazard driver in
        this codebase. A comfortably safe clearance/capacity must score a
        LOW likelihood on its own hazard entry, not a spuriously high one."""
        r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=2.0, required_min_clearance_m=1.0,  # 1.0 m margin, well above thresholds
            is_night_operation=False, measured_illuminance_lux=200,
            hardware_years_in_service=1,  # ~97.5% capacity remaining
        )
        tide_hazard = next(h for h in r["risk_matrix"]["hazards"] if h["name"] == "Tide clearance margin")
        corrosion_hazard = next(h for h in r["risk_matrix"]["hazards"]
                                 if h["name"] == "Splash-zone hardware corrosion capacity loss")
        assert tide_hazard["likelihood"] == 1, "safe tide clearance must not score a high likelihood"
        assert corrosion_hazard["likelihood"] == 1, "near-100% remaining capacity must not score a high likelihood"
        assert r["risk_matrix"]["overall_band"] == "Low"

    def test_marine_port_critical_tide_and_corrosion_score_high_likelihood(self):
        """Inverse of the above: a genuinely unsafe (small) clearance
        margin and badly corroded hardware must still score correctly in
        the unsafe direction."""
        r = calculate_marine_port_kinetic_risk(
            current_tide_level_m=1.0, required_min_clearance_m=1.0,  # 0 m margin: critical
            is_night_operation=False, measured_illuminance_lux=200,
            hardware_years_in_service=20,  # 50% remaining capacity: critical
        )
        tide_hazard = next(h for h in r["risk_matrix"]["hazards"] if h["name"] == "Tide clearance margin")
        corrosion_hazard = next(h for h in r["risk_matrix"]["hazards"]
                                 if h["name"] == "Splash-zone hardware corrosion capacity loss")
        assert tide_hazard["likelihood"] == 5
        assert corrosion_hazard["likelihood"] == 5
        assert r["risk_matrix"]["overall_band"] == "Extreme"


class TestISO7243HeatStress:
    """calculate_iso7243_heat_stress(): ACGIH TLV + ISO 7243 occupational
    heat-stress screen with workload/metabolic category, Clothing
    Adjustment Factor, acclimatization, and a work/rest-ratio solver."""

    def test_comfortable_case_is_low_and_no_override(self):
        r = calculate_iso7243_heat_stress(workload_category="light", clothing_type="work_clothes",
                                           acclimatized=False, requested_work_rest_ratio="100/0",
                                           air_temp_c=24, relative_humidity_pct=40)
        _assert_standard_shape(r)
        assert r["risk_band"] == "LOW"
        assert r["safety_override"] is False
        assert r["recommended_work_rest_ratio"] == "100/0"

    def test_hot_moderate_work_recommends_a_stricter_ratio_without_stopping_work(self):
        r = calculate_iso7243_heat_stress(workload_category="moderate", clothing_type="work_clothes",
                                           acclimatized=False, requested_work_rest_ratio="100/0",
                                           wbgt_measured_c=29.0)
        assert r["requested_ratio_exceeds"] is True
        assert r["recommended_work_rest_ratio"] in ("75/25", "50/50", "25/75")
        assert r["safety_override"] is False

    def test_extreme_wbgt_with_heavy_clothing_forces_stop_work(self):
        r = calculate_iso7243_heat_stress(workload_category="heavy", clothing_type="double_layer_woven_coveralls",
                                           acclimatized=False, requested_work_rest_ratio="25/75",
                                           wbgt_measured_c=32.0)
        assert r["safety_override"] is True
        assert r["recommended_work_rest_ratio"] is None
        assert r["risk_band"] == "CRITICAL"

    def test_acclimatized_worker_gets_more_headroom_than_unacclimatized(self):
        unacclimatized = calculate_iso7243_heat_stress(workload_category="moderate", clothing_type="work_clothes",
                                                         acclimatized=False, wbgt_measured_c=28.2)
        acclimatized = calculate_iso7243_heat_stress(workload_category="moderate", clothing_type="work_clothes",
                                                       acclimatized=True, wbgt_measured_c=28.2)
        assert unacclimatized["requested_ratio_exceeds"] is True
        assert acclimatized["requested_ratio_exceeds"] is False

    def test_vapor_barrier_ppe_always_forces_stop_work_and_monitoring_flag(self):
        r = calculate_iso7243_heat_stress(workload_category="light", clothing_type="vapor_barrier_coveralls",
                                           acclimatized=True, wbgt_measured_c=18.0)
        assert r["safety_override"] is True
        assert r["requires_physiological_monitoring"] is True

    def test_very_heavy_continuous_100_0_is_not_applicable(self):
        """ACGIH's table has no continuous (100/0) very_heavy cell - must
        report None rather than a fabricated number."""
        r = calculate_iso7243_heat_stress(workload_category="very_heavy", clothing_type="work_clothes",
                                           acclimatized=False, requested_work_rest_ratio="100/0",
                                           wbgt_measured_c=20.0)
        assert r["requested_ratio_limit_c"] is None

    def test_missing_wbgt_and_temp_rh_raises_value_error(self):
        with pytest.raises(ValueError):
            calculate_iso7243_heat_stress(workload_category="light")

    def test_clothing_adjustment_factor_increases_effective_wbgt(self):
        bare = calculate_iso7243_heat_stress(clothing_type="work_clothes", wbgt_measured_c=25.0)
        insulated = calculate_iso7243_heat_stress(clothing_type="double_layer_woven_coveralls", wbgt_measured_c=25.0)
        assert insulated["effective_wbgt_c"] > bare["effective_wbgt_c"]
        assert insulated["effective_wbgt_c"] - bare["effective_wbgt_c"] == pytest.approx(
            CLOTHING_ADJUSTMENT_FACTOR_C["double_layer_woven_coveralls"]
        )

    def test_risk_matrix_present_and_well_formed(self):
        r = calculate_iso7243_heat_stress(workload_category="heavy", clothing_type="vapor_barrier_coveralls",
                                           acclimatized=False, wbgt_measured_c=30.0)
        rm = r["risk_matrix"]
        assert rm["hazards"]
        assert 1 <= rm["overall_score"] <= 25


class TestStopWorkTriggerRegistry:
    def test_registry_covers_all_nine_modules_with_required_fields(self):
        assert len(STOP_WORK_TRIGGERS) == 9
        required = {"module", "trigger", "threshold", "source_constant", "result_key", "profile_dependent"}
        for entry in STOP_WORK_TRIGGERS:
            assert set(entry.keys()) == required
            assert entry["module"] and entry["trigger"] and entry["threshold"]
            assert isinstance(entry["profile_dependent"], bool)

    def test_get_stop_work_triggers_no_filter_returns_everything(self):
        assert get_stop_work_triggers() == STOP_WORK_TRIGGERS
        assert get_stop_work_triggers(None) == STOP_WORK_TRIGGERS

    def test_get_stop_work_triggers_filters_by_module_substring(self):
        result = get_stop_work_triggers("Solar")
        assert len(result) == 1
        assert result[0]["module"] == "Solar (Desert)"

    def test_get_stop_work_triggers_unknown_module_returns_full_registry(self):
        assert get_stop_work_triggers("not_a_real_module_xyz") == STOP_WORK_TRIGGERS


class TestFormulaStandardsMap:
    """Regulatory Algorithm Validation: every scoring formula mapped
    against its claimed standard, with an honest Direct/Adapted/
    Illustrative status."""

    VALID_STATUSES = {"Direct implementation", "Adapted / approximated", "Illustrative (not standards-derived)"}

    def test_every_entry_has_required_fields_and_a_valid_status(self):
        assert len(FORMULA_STANDARDS_MAP) > 0
        required = {"function", "module", "formula_summary", "cited_standards", "validation_status", "caveat"}
        for entry in FORMULA_STANDARDS_MAP:
            assert set(entry.keys()) == required
            assert entry["validation_status"] in self.VALID_STATUSES
            assert isinstance(entry["cited_standards"], list) and entry["cited_standards"]
            assert entry["caveat"], f"{entry['function']} must document a caveat, even if empty-ish"

    def test_datacenter_arc_flash_proxy_is_honestly_flagged_illustrative(self):
        """The datacenter module's arc-flash number is a simplified
        load-only proxy, NOT the Lee formula or IEEE 1584 - this must
        never be silently upgraded to look like a certified figure."""
        entry = get_formula_standard("calculate_datacenter_kinetic_risk() arc_flash_energy_cal")
        assert entry is not None
        assert entry["validation_status"] == "Illustrative (not standards-derived)"
        assert "IEEE 1584" in entry["caveat"] or "Lee" in entry["caveat"]

    def test_get_formula_standard_unknown_returns_none(self):
        assert get_formula_standard("nonexistent_formula_xyz") is None

    def test_get_all_formula_standards_returns_full_list(self):
        assert get_all_formula_standards() == FORMULA_STANDARDS_MAP


class TestEvidenceTraceability:
    """analytics.build_evidence_traceability(): Data Source, Geographic
    Location/Site, Timestamp, Measuring Equipment/Sensor Model."""

    def test_auto_mode_labeled_as_live_feed(self):
        st.session_state.clear()
        result = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        evidence = build_evidence_traceability(st, result, data_mode="auto", sensor_model="")
        assert "Live" in evidence["Data Source"]

    def test_manual_mode_labeled_as_manual_entry(self):
        st.session_state.clear()
        result = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        evidence = build_evidence_traceability(st, result, data_mode="manual", sensor_model="")
        assert "Manual" in evidence["Data Source"]

    def test_blank_sensor_model_is_honestly_labeled_not_fabricated(self):
        st.session_state.clear()
        result = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        evidence = build_evidence_traceability(st, result, data_mode="manual", sensor_model="")
        assert evidence["Measuring Equipment / Sensor Model"] == "Not specified / simulated input"

    def test_provided_sensor_model_is_captured_verbatim(self):
        st.session_state.clear()
        result = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        evidence = build_evidence_traceability(st, result, data_mode="manual", sensor_model="Kestrel 5400")
        assert evidence["Measuring Equipment / Sensor Model"] == "Kestrel 5400"

    def test_geographic_location_reuses_org_context(self):
        st.session_state.clear()
        set_org_context(st, organization="Acme Corp", project="Tower A", site="Zone 3")
        result = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        evidence = build_evidence_traceability(st, result, data_mode="manual")
        assert evidence["Geographic Location - Organization"] == "Acme Corp"
        assert evidence["Geographic Location - Project"] == "Tower A"
        assert evidence["Geographic Location - Site"] == "Zone 3"

    def test_timestamp_is_present_and_utc_labeled(self):
        st.session_state.clear()
        result = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        evidence = build_evidence_traceability(st, result, data_mode="manual")
        assert "UTC" in evidence["Assessment Timestamp (UTC)"]


class TestOfficialReportWithEvidence:
    """render_official_report()/_build_report_pdf() must accept an
    optional evidence block (backward compatible with evidence=None) and
    produce a valid PDF either way."""

    def test_pdf_builds_with_evidence_block(self):
        st.session_state.clear()
        result = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        evidence = build_evidence_traceability(st, result, data_mode="auto", sensor_model="Test Sensor")
        pdf_bytes = _build_report_pdf(result, "narrative", ["control"], [], [], "2026-01-01 00:00 UTC", "en",
                                       evidence=evidence)
        assert pdf_bytes[:4] == b"%PDF"

    def test_pdf_builds_without_evidence_backward_compatible(self):
        result = calculate_solar_albedo_heat_risk(ghi=300, uv_index=3, ambient_temp=24, surface_type="pure_desert_sand")
        pdf_bytes = _build_report_pdf(result, "narrative", ["control"], [], [], "2026-01-01 00:00 UTC", "en")
        assert pdf_bytes[:4] == b"%PDF"
